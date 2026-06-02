"""Tests for the Yardi AR Aging file parser."""

import pytest
import openpyxl

from app.models import ARAgingRow
from app.parser.ar_aging import parse_ar_aging_reports


def _make_ar_wb(tmp_path, filename, rows):
    """Helper: create a synthetic Yardi AR Aging workbook."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report1"
    ws.append(["Affordable Aging Detail"])
    ws.append(["Property: Affirmed Property List (affirmed)"])
    period = f"Post To(MM/YY): {filename.split('_')[-2].zfill(2)}/{filename.split('_')[-1].split('.')[0]}"
    ws.append([period])
    ws.append(["Property Name", "Charge Amount", "Current Owed",
               "0-30 Owed", "31-60 Owed", "61-90 Owed",
               "Over 90 Owed", "Pre-payments", "Suspense"])
    ws.append([""] * 9)
    for r in rows:
        ws.append(r)
    ws.append([None] * 9)
    ws.append(["Grand Total"] + [0] * 8)
    path = tmp_path / filename
    wb.save(str(path))
    return str(path)


@pytest.fixture
def tenant_rent_wb(tmp_path):
    return _make_ar_wb(tmp_path, "Solari_AR Aging_Tenant Rent_03_2024.xlsx", [
        ["Alora Family (alora)",  10000, 8000, 500, 300, 200, 100, -50,   0],
        ["Beechwood (beech)",     20000, 15000, 1000, 800, 600, 400, -100, 0],
    ])


@pytest.fixture
def subsidy_wb(tmp_path):
    return _make_ar_wb(tmp_path, "Solari_AR Aging_Subsidy_03_2024.xlsx", [
        ["Alora Family (alora)", 5000, 4000, 200, 100, 50, 25, 0, 0],
    ])


@pytest.fixture
def tenant_receivable_wb(tmp_path):
    return _make_ar_wb(tmp_path, "ConAm_AR Aging_Tenant Receivable_06_2024.xlsx", [
        ["Alora Family (alora)", 10000, 8000, 500, 300, 200, 100, -50, 0],
    ])


@pytest.fixture
def subsidy_receivable_wb(tmp_path):
    return _make_ar_wb(tmp_path, "ConAm_AR Aging_Subsidy Receivable_06_2024.xlsx", [
        ["Alora Family (alora)", 5000, 4000, 200, 100, 50, 25, 0, 0],
    ])


def test_returns_ar_aging_row_instances(tenant_rent_wb):
    rows = parse_ar_aging_reports([tenant_rent_wb])
    assert len(rows) == 2
    assert all(isinstance(r, ARAgingRow) for r in rows)


def test_receivable_type_tenant_rent(tenant_rent_wb):
    rows = parse_ar_aging_reports([tenant_rent_wb])
    assert all(r.receivable_type == "Tenant Rent" for r in rows)


def test_receivable_type_subsidy(subsidy_wb):
    rows = parse_ar_aging_reports([subsidy_wb])
    assert all(r.receivable_type == "Subsidy" for r in rows)


def test_type_normalization_tenant_receivable(tenant_receivable_wb):
    rows = parse_ar_aging_reports([tenant_receivable_wb])
    assert all(r.receivable_type == "Tenant Rent" for r in rows)


def test_type_normalization_subsidy_receivable(subsidy_receivable_wb):
    rows = parse_ar_aging_reports([subsidy_receivable_wb])
    assert all(r.receivable_type == "Subsidy" for r in rows)


def test_pm_name_from_filename(tenant_rent_wb):
    rows = parse_ar_aging_reports([tenant_rent_wb])
    assert all(r.pm_name == "Solari" for r in rows)


def test_year_month_from_filename(tenant_rent_wb):
    rows = parse_ar_aging_reports([tenant_rent_wb])
    assert all(r.year == 2024 and r.month == 3 for r in rows)


def test_property_code_suffix_stripped(tenant_rent_wb):
    rows = parse_ar_aging_reports([tenant_rent_wb])
    prop_names = {r.property_name for r in rows}
    assert "Alora Family" in prop_names
    assert "Beechwood" in prop_names
    assert not any("(" in name for name in prop_names)


def test_numeric_fields(tenant_rent_wb):
    rows = parse_ar_aging_reports([tenant_rent_wb])
    alora = next(r for r in rows if "Alora" in r.property_name)
    assert alora.charge_amount  == pytest.approx(10000)
    assert alora.current_owed   == pytest.approx(8000)
    assert alora.owed_0_30      == pytest.approx(500)
    assert alora.owed_31_60     == pytest.approx(300)
    assert alora.owed_61_90     == pytest.approx(200)
    assert alora.owed_over_90   == pytest.approx(100)
    assert alora.prepayments    == pytest.approx(-50)


def test_grand_total_row_excluded(tenant_rent_wb):
    rows = parse_ar_aging_reports([tenant_rent_wb])
    assert not any("Grand Total" in r.property_name for r in rows)


def test_blank_row_stops_iteration(tenant_rent_wb):
    rows = parse_ar_aging_reports([tenant_rent_wb])
    assert all(r.property_name for r in rows)


def test_total_overdue_computed(tenant_rent_wb):
    rows = parse_ar_aging_reports([tenant_rent_wb])
    alora = next(r for r in rows if "Alora" in r.property_name)
    # total_over_60 = owed_61_90 + owed_over_90 = 200 + 100 = 300 (excludes 31-60 bucket)
    assert alora.total_overdue == pytest.approx(300)


def test_pct_overdue_computed(tenant_rent_wb):
    rows = parse_ar_aging_reports([tenant_rent_wb])
    alora = next(r for r in rows if "Alora" in r.property_name)
    # pct_over_60 = 300 / 10000 = 0.03 (only 61+ days past due)
    assert alora.pct_overdue == pytest.approx(0.03)


def test_none_numeric_treated_as_zero(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report1"
    ws.append(["Affordable Aging Detail"])
    ws.append(["Property: Affirmed Property List (affirmed)"])
    ws.append(["Post To(MM/YY): 03/2024"])
    ws.append(["Property Name", "Charge Amount", "Current Owed",
               "0-30 Owed", "31-60 Owed", "61-90 Owed",
               "Over 90 Owed", "Pre-payments", "Suspense"])
    ws.append([""] * 9)
    ws.append(["Test Prop (test)", 5000, 5000, None, None, None, None, None, 0])
    ws.append([None] * 9)
    ws.append(["Grand Total"] + [0] * 8)
    path = tmp_path / "PM_AR Aging_Subsidy_03_2024.xlsx"
    wb.save(str(path))
    rows = parse_ar_aging_reports([str(path)])
    assert len(rows) == 1
    assert rows[0].owed_31_60   == 0.0
    assert rows[0].prepayments  == 0.0
    assert rows[0].pct_overdue  == 0.0


def test_source_file_is_basename(tenant_rent_wb):
    rows = parse_ar_aging_reports([tenant_rent_wb])
    assert all(r.source_file == "Solari_AR Aging_Tenant Rent_03_2024.xlsx" for r in rows)


def test_multiple_files_combined(tenant_rent_wb, subsidy_wb):
    rows = parse_ar_aging_reports([tenant_rent_wb, subsidy_wb])
    types = {r.receivable_type for r in rows}
    assert types == {"Tenant Rent", "Subsidy"}


# ─── ConAm 2026+ format tests ────────────────────────────────────────────────

def _make_conam_2026_wb(tmp_path, filename="ConAm_AR Aging_03_2026.xlsx"):
    """Minimal ConAm 2026+ format workbook."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report1"
    ws.append(["Affordable Receivable Aging Summary"])           # row 1
    ws.append(["Property: Client Affirmed List (.claffir)"])     # row 2
    ws.append(["Trans through: 03/2026"])                        # row 3
    ws.append([None, None, "Total"])                             # row 4
    ws.append(["Charge", None, "Unpaid", "0-30", "31-60", "61-90", "Over 90", None, None, None])  # row 5
    ws.append(["Code", "Description", "Charges", "days", "days", "days", "days", "Prepays", "Suspense", "Balance"])  # row 6
    # Property block 1: "Link, The (f66)"
    ws.append(["Link, The (f66)", None, None, None, None, None, None, None, None, None])
    ws.append(["RENT",    "Rent",            50000, 10000, 30000, 0, 0,     0,     0, 50000])
    ws.append(["RENTPM",  "Prior Month Rent",  500,     0,   500, 0, 0,     0,     0,   500])
    ws.append(["SUBRNT",  "Subsidy Rent CR", 20000,  5000, 15000, 0, 0,     0,     0, 20000])
    ws.append(["KEYS",    "Keys/Locks/Gate",    10,     0,    10, 0, 0,     0,     0,    10])
    ws.append(["Total",   None,              70510, 15000, 45510, 0, 0,     0,     0, 70510])
    # Property block 2: "Cypress (f78)"
    ws.append([None])
    ws.append(["Cypress (f78)", None, None, None, None, None, None, None, None, None])
    ws.append(["RENT",   "Rent",             8000,  2000,  5000, 500, 500,  0,   0, 8000])
    ws.append(["SUBRNT", "Subsidy Rent CR",  3000,  1000,  1500,   0, 500, -100, 0, 2900])
    ws.append(["Total",  None,              11000,  3000,  6500, 500,1000, -100,  0,10900])
    ws.append([None])
    ws.append(["Grand Total", None, 81510, 18000, 52010, 500, 1000, -100, 0, 81410])
    path = tmp_path / filename
    wb.save(str(path))
    return str(path)


@pytest.fixture
def conam_2026_wb(tmp_path):
    return _make_conam_2026_wb(tmp_path)


def test_conam_2026_format_detected(conam_2026_wb):
    rows = parse_ar_aging_reports([conam_2026_wb])
    assert len(rows) > 0


def test_conam_2026_both_types_extracted(conam_2026_wb):
    rows = parse_ar_aging_reports([conam_2026_wb])
    types = {r.receivable_type for r in rows}
    assert types == {"Tenant Rent", "Subsidy"}


def test_conam_2026_property_names(conam_2026_wb):
    rows = parse_ar_aging_reports([conam_2026_wb])
    props = {r.property_name for r in rows}
    # "Link, The (f66)" → stripped code → fixed inversion → "The Link"
    assert "The Link" in props
    assert "Cypress" in props
    assert not any("(" in p for p in props)


def test_conam_2026_year_month(conam_2026_wb):
    rows = parse_ar_aging_reports([conam_2026_wb])
    assert all(r.year == 2026 and r.month == 3 for r in rows)


def test_conam_2026_pm_name(conam_2026_wb):
    rows = parse_ar_aging_reports([conam_2026_wb])
    assert all(r.pm_name == "ConAm" for r in rows)


def test_conam_2026_rent_and_rentpm_summed(conam_2026_wb):
    rows = parse_ar_aging_reports([conam_2026_wb])
    link_rent = next(r for r in rows if r.property_name == "The Link"
                     and r.receivable_type == "Tenant Rent")
    # RENT(50000) + RENTPM(500) = 50500
    assert link_rent.charge_amount == pytest.approx(50500)
    assert link_rent.current_owed  == pytest.approx(50500)


def test_conam_2026_subsidy_amounts(conam_2026_wb):
    rows = parse_ar_aging_reports([conam_2026_wb])
    link_sub = next(r for r in rows if r.property_name == "The Link"
                    and r.receivable_type == "Subsidy")
    assert link_sub.charge_amount == pytest.approx(20000)
    assert link_sub.owed_0_30     == pytest.approx(5000)
    assert link_sub.owed_31_60    == pytest.approx(15000)


def test_conam_2026_non_rent_codes_excluded(conam_2026_wb):
    """KEYS and similar charge codes must not produce ARAgingRow records."""
    rows = parse_ar_aging_reports([conam_2026_wb])
    # Only RENT/RENTPM and SUBRNT per property → 4 rows total (2 props × 2 types)
    assert len(rows) == 4


# ─── ConAm 2025 format tests ─────────────────────────────────────────────────

def _make_conam_2025_wb(tmp_path, filename="ConAm_AR Aging_12_2025.xlsx"):
    """Minimal ConAm 2025 format workbook (two property sheets)."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default sheet

    for prop, rent_bal, sub_bal in [
        ("Link",   [27599, 25090, 0, 22535, 139659, 214883],
                   [10239,  6592, 0,  4425, 242799, 264055]),
        ("Cypress",[5856,   4353, 0,  3285,  12634,  26128],
                   [2475,   2500, 0,  2298, 117358, 124631]),
    ]:
        ws = wb.create_sheet(prop)
        ws.cell(1, 1, "A/R Aging Report")
        # Row 4 (index 3): filter criteria with "As of 12/31/25"
        ws.cell(4, 1, "Community = X\nAs of 12/31/25\nSubsidy = INCLUDE")
        # "Community Totals By Balance Type:" marker in col B (index 1)
        ws.cell(26, 2, "Community Totals By Balance Type:")
        ws.cell(26, 23, "Totals")   # col W (0-based 22 = 1-based 23)
        # Header row
        ws.cell(31, 23, "Current")
        ws.cell(31, 26, "Over 30")
        ws.cell(31, 29, "Over 60")
        ws.cell(31, 32, "Over 90")
        ws.cell(31, 35, "Over 120")
        ws.cell(31, 37, "Balance")
        # "Rent -" balance type row — col R (1-based 18 = 0-based 17)
        ws.cell(33, 18, "Rent -")
        ws.cell(33, 23, rent_bal[0]);  ws.cell(33, 26, rent_bal[1])
        ws.cell(33, 29, rent_bal[2]);  ws.cell(33, 32, rent_bal[3])
        ws.cell(33, 35, rent_bal[4]);  ws.cell(33, 37, rent_bal[5])
        # "Subsidy - HASanDiego" balance type row
        ws.cell(34, 18, "Subsidy - HASanDiego")
        ws.cell(34, 23, sub_bal[0]);  ws.cell(34, 26, sub_bal[1])
        ws.cell(34, 29, sub_bal[2]);  ws.cell(34, 32, sub_bal[3])
        ws.cell(34, 35, sub_bal[4]);  ws.cell(34, 37, sub_bal[5])
        # "Rent Concession -" — should be excluded
        ws.cell(37, 18, "Rent Concession -")
        ws.cell(37, 23, -100)

    path = tmp_path / filename
    wb.save(str(path))
    return str(path)


@pytest.fixture
def conam_2025_wb(tmp_path):
    return _make_conam_2025_wb(tmp_path)


def test_conam_2025_format_detected(conam_2025_wb):
    rows = parse_ar_aging_reports([conam_2025_wb])
    assert len(rows) > 0


def test_conam_2025_both_types_per_property(conam_2025_wb):
    rows = parse_ar_aging_reports([conam_2025_wb])
    # 2 properties × 2 types = 4 rows
    assert len(rows) == 4
    types = {r.receivable_type for r in rows}
    assert types == {"Tenant Rent", "Subsidy"}


def test_conam_2025_property_names(conam_2025_wb):
    rows = parse_ar_aging_reports([conam_2025_wb])
    props = {r.property_name for r in rows}
    assert "Link" in props    # PROPERTY_NAME_MAP → "The Link" in upload pipeline
    assert "Cypress" in props


def test_conam_2025_year_month_from_filename(conam_2025_wb):
    rows = parse_ar_aging_reports([conam_2025_wb])
    assert all(r.year == 2025 and r.month == 12 for r in rows)


def test_conam_2025_year_month_fallback(tmp_path):
    """When filename lacks period, parser falls back to 'As of' cell."""
    wb_path = _make_conam_2025_wb(tmp_path, "ConAm_AR Aging_no_period.xlsx")
    rows = parse_ar_aging_reports([wb_path])
    # Sheet contains "As of 12/31/25" → year=2025, month=12
    assert all(r.year == 2025 and r.month == 12 for r in rows)


def test_conam_2025_over120_merged_into_over90(conam_2025_wb):
    rows = parse_ar_aging_reports([conam_2025_wb])
    link_rent = next(r for r in rows if r.property_name == "Link"
                     and r.receivable_type == "Tenant Rent")
    # over_90=22535, over_120=139659 → owed_over_90=162194
    assert link_rent.owed_over_90 == pytest.approx(22535 + 139659)


def test_conam_2025_rent_concession_excluded(conam_2025_wb):
    """'Rent Concession -' balance type must NOT be treated as Tenant Rent."""
    rows = parse_ar_aging_reports([conam_2025_wb])
    for r in rows:
        # The -100 concession row should not inflate any Tenant Rent balance
        if r.property_name == "Link" and r.receivable_type == "Tenant Rent":
            # owed_0_30 = 27599 (Rent only, not -100 from concession)
            assert r.owed_0_30 == pytest.approx(27599)


def test_conam_2025_subsidy_amounts(conam_2025_wb):
    rows = parse_ar_aging_reports([conam_2025_wb])
    link_sub = next(r for r in rows if r.property_name == "Link"
                    and r.receivable_type == "Subsidy")
    assert link_sub.owed_0_30  == pytest.approx(10239)
    assert link_sub.owed_31_60 == pytest.approx(6592)
