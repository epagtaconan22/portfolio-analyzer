"""Shared openpyxl styles, number formats, and KPI header comment helper."""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from config import KPI_FORMULAS

# Colors
GREEN_FILL  = PatternFill("solid", fgColor="C6EFCE")
RED_FILL    = PatternFill("solid", fgColor="FFC7CE")
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
SUBHDR_FILL = PatternFill("solid", fgColor="2E75B6")

HEADER_FONT     = Font(bold=True, color="FFFFFF", size=11)
SUBHEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
BOLD_FONT       = Font(bold=True)
DEFAULT_FONT    = Font(size=10)

CURRENCY_FMT = '$#,##0;[Red]($#,##0)'
PCT_FMT      = '0.0%'
VAR_PCT_FMT  = '0.0%;[Red](0.0%)'
COMMA_FMT    = '#,##0'

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

def style_header_row(ws, row_num: int, num_cols: int, fill=HEADER_FILL, font=HEADER_FONT):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def add_kpi_comment(cell, kpi_name: str, author: str = "Portfolio Analyzer"):
    formula = KPI_FORMULAS.get(kpi_name)
    if formula:
        comment = Comment(formula, author)
        comment.width = 300
        comment.height = 100
        cell.comment = comment


def apply_variance_fill(cell, value, favorable_is_positive: bool = True,
                        threshold: float = 0.0):
    """Green if favorable, red if unfavorable.

    threshold (default 0): abs(value) must EXCEED threshold before any fill is applied.
    Pass threshold=0.05 to only colour cells where the variance exceeds ±5%.
    """
    if value is None:
        return
    if abs(value) <= threshold:
        return  # within the neutral band — no highlight
    if favorable_is_positive:
        cell.fill = GREEN_FILL if value >= 0 else RED_FILL
    else:
        cell.fill = GREEN_FILL if value <= 0 else RED_FILL
