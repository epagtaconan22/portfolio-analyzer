"""Builds the 4-tab presentation workbook: Dashboard, Property Analysis, Property Monthly KPIs, AR Aging."""

import os
from typing import Optional
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.comments import Comment
from openpyxl.worksheet.table import Table, TableStyleInfo
from app.models import PropertyPeriodKPIs
from app.exporter.styles import (
    style_header_row, add_kpi_comment, apply_variance_fill,
    CURRENCY_FMT, PCT_FMT, VAR_PCT_FMT, COMMA_FMT, BOLD_FONT,
    SUBHDR_FILL, SUBHEADER_FONT,
)
from config import ECO_OCC_TARGET, KPI_FORMULAS

_GROUP_PARENTS = {"Actual Income", "Actual Expenses", "Actual NOI", "GPR",
                  "Eco Occ %", "Physical Occ %", "Leakage Gap"}

# Section header style: darker navy (#1F4E79) matching the app's primary colour
_DARK_HDR_FILL = PatternFill("solid", fgColor="1F4E79")
_DARK_HDR_FONT = Font(bold=True, color="FFFFFF", size=10)

# Alternating table row fill (ice blue)
_ICE_BLUE_FILL = PatternFill("solid", fgColor="DEEAF1")
_WHITE_FILL    = PatternFill("solid", fgColor="FFFFFF")

# Traffic-light fills for Leakage Gap
_LEAKAGE_GREEN  = PatternFill("solid", fgColor="C6EFCE")   # 0 – 2%  healthy
_LEAKAGE_YELLOW = PatternFill("solid", fgColor="FFEB9C")   # 3 – 5%  caution
_LEAKAGE_RED    = PatternFill("solid", fgColor="FFC7CE")   # > 5%    problem

_LEAKAGE_GAP_NOTE = (
    "Leakage Gap = Physical Occ % − Economic Occ %\n\n"
    "Interpretation guide:\n"
    "  0–2%: Very healthy. Minimal leakage. Usually normal timing or small resident balances.\n"
    "  2–3%: Still generally healthy if stable and explained. Watch trend, but not automatically a concern.\n"
    "  3–5%: Caution range. Requires explanation: A/R aging, concessions, subsidy delays, "
    "vacancy loss, bad debt, or posting issues.\n"
    "  5% or greater: Problem range. High physical occupancy may be masking collection weakness "
    "or revenue leakage."
)

_AR_MONTH_ABBR = {
    1: "Jan", 2: "Feb", 3: "Mar", 4: "Apr", 5: "May", 6: "Jun",
    7: "Jul", 8: "Aug", 9: "Sep", 10: "Oct", 11: "Nov", 12: "Dec",
}

# Budget-based YoY comparison: maps each KPI key to the field looked up when
# computing "current year budget vs prior year budget."  Keys absent → blank cells.
_BUDGET_YOY_KEY: dict[str, str] = {
    "actual_income":        "budget_income",
    "budget_income":        "budget_income",
    "income_variance":      "income_variance",
    "income_variance_pct":  "income_variance_pct",
    "actual_expenses":      "budget_expenses",
    "budget_expenses":      "budget_expenses",
    "expense_variance":     "expense_variance",
    "expense_variance_pct": "expense_variance_pct",
    "actual_noi":           "budget_noi",
    "budget_noi":           "budget_noi",
    "noi_variance":         "noi_variance",
    "noi_variance_pct":     "noi_variance_pct",
    "eco_occ_pct":          "budget_eco_occ_pct",
}

# Currency keys get both a Δ$ column and a Δ% column; pct keys get Δpp only.
_YOY_CURRENCY_KEYS = frozenset({
    "actual_income", "budget_income", "income_variance",
    "actual_expenses", "budget_expenses", "expense_variance",
    "actual_noi", "budget_noi", "noi_variance",
})
_YOY_PCT_KEYS = frozenset({
    "income_variance_pct", "expense_variance_pct", "noi_variance_pct", "eco_occ_pct",
})
_YOY_FAVORABLE_IF_POSITIVE = frozenset({
    "actual_income", "budget_income",
    "actual_noi",    "budget_noi",    "noi_variance",
    "eco_occ_pct",
})


def _ar_period_label(year: int, month: int) -> str:
    return f"{_AR_MONTH_ABBR.get(month, str(month))}-{year}"


# ─── Quarter helpers ─────────────────────────────────────────────────────────

def _month_to_quarter(month: int) -> int:
    """Return quarter number (1–4) for a given month (1–12)."""
    return (month - 1) // 3 + 1


def _quarter_label(year: int, quarter: int) -> str:
    """Return display label, e.g. 'Q1 - 2025'."""
    return f"Q{quarter} - {year}"


def _get_sorted_quarters(kpis) -> list[tuple[int, int]]:
    """Return (year, quarter) tuples present in the KPI list, newest first."""
    quarters: set[tuple[int, int]] = set()
    for k in kpis:
        quarters.add((k.year, _month_to_quarter(k.month)))
    return sorted(quarters, reverse=True)


def _kpis_for_quarter(kpis, year: int, quarter: int) -> list:
    """Return KPI records belonging to the given year and quarter."""
    months = {(quarter - 1) * 3 + 1, (quarter - 1) * 3 + 2, (quarter - 1) * 3 + 3}
    return [k for k in kpis if k.year == year and k.month in months]


# ─── Dashboard KPI row definitions ───────────────────────────────────────────
# Each entry: (display_label, _aggregate()-dict-key, number_format)
# None entries produce a blank separator row in the transposed table.

_DASHBOARD_KPI_ROWS = [
    ("Actual Income",       "actual_income",        CURRENCY_FMT),
    ("Budget Income",       "budget_income",        CURRENCY_FMT),
    ("Income Variance",     "income_variance",      CURRENCY_FMT),
    ("Income Variance %",   "income_variance_pct",  PCT_FMT),
    None,
    ("Actual Expenses",     "actual_expenses",      CURRENCY_FMT),
    ("Budget Expenses",     "budget_expenses",      CURRENCY_FMT),
    ("Expense Variance",    "expense_variance",     CURRENCY_FMT),
    ("Expense Variance %",  "expense_variance_pct", PCT_FMT),
    None,
    ("Actual NOI",          "actual_noi",           CURRENCY_FMT),
    ("Budget NOI",          "budget_noi",           CURRENCY_FMT),
    ("NOI Variance",        "noi_variance",         CURRENCY_FMT),
    ("NOI Variance %",      "noi_variance_pct",     PCT_FMT),
    None,
    ("GPR",                 "gpr",                  CURRENCY_FMT),
    ("Vacancy",             "vacancy",              CURRENCY_FMT),
    ("Concessions",         "concessions",          CURRENCY_FMT),
    ("Bad Debt",            "bad_debt",             CURRENCY_FMT),
    ("Net Collectible",     "net_collectible",      CURRENCY_FMT),
    ("Eco Occ %",           "eco_occ_pct",          PCT_FMT),
    ("Budget Eco Occ %",    "budget_eco_occ_pct",   PCT_FMT),
    ("Eco Occ Variance",    "eco_occ_variance",     PCT_FMT),
    ("Physical Occ %",      "physical_occ_pct",     PCT_FMT),
    ("Leakage Gap",         "leakage_gap",          PCT_FMT),
    None,
    ("Income/Unit",         "income_per_unit",      CURRENCY_FMT),
    ("Expense/Unit",        "expense_per_unit",     CURRENCY_FMT),
    ("NOI/Unit",            "noi_per_unit",         CURRENCY_FMT),
]


# ─── Public entry point ───────────────────────────────────────────────────────

def build_main_workbook(
    kpis: list[PropertyPeriodKPIs],
    portfolio_name: str,
    output_path: str,
    eco_occ_target: float = ECO_OCC_TARGET,
    ar_rows: list | None = None,
    use_budget_eco_occ: bool = False,
) -> str:
    """Builds main workbook at output_path. Returns path."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _build_dashboard(wb, kpis, portfolio_name, eco_occ_target, ar_rows=ar_rows,
                     use_budget_eco_occ=use_budget_eco_occ)
    _build_property_analysis(wb, kpis, portfolio_name, eco_occ_target)
    _build_monthly_kpis(wb, kpis)
    _build_ar_aging(wb, ar_rows or [], portfolio_name, kpis=kpis)

    wb.save(output_path)
    return output_path


# ─── Dashboard ────────────────────────────────────────────────────────────────

def _build_dashboard(wb, kpis, portfolio_name, eco_occ_target, ar_rows=None,
                     use_budget_eco_occ=False):
    ws = wb.create_sheet("Dashboard")

    from openpyxl.worksheet.properties import Outline
    ws.sheet_properties.outlinePr = Outline(summaryBelow=False, summaryRight=False)

    props = {k.property_name for k in kpis if not k.is_carveout and not k.is_partial_year}
    num_props = len(props)

    # ── Quarter-period aggregates (partial-year properties excluded) ─────────
    quarters = _get_sorted_quarters(kpis)
    period_labels = [_quarter_label(yr, q) for (yr, q) in quarters]
    period_aggs: dict[str, dict] = {}
    for (yr, q) in quarters:
        q_kpis = [k for k in _kpis_for_quarter(kpis, yr, q)
                  if not k.is_carveout and not k.is_partial_year]
        period_aggs[_quarter_label(yr, q)] = _aggregate(q_kpis)

    # ── Annual aggregates and YoY year pairs (partial-year excluded) ─────────
    years_sorted = sorted({k.year for k in kpis if not k.is_carveout and not k.is_partial_year})
    year_aggs: dict[int, dict] = {}
    for yr in years_sorted:
        yr_kpis = [k for k in kpis
                   if k.year == yr and not k.is_carveout and not k.is_partial_year]
        year_aggs[yr] = _aggregate(yr_kpis)
    # One year pair per consecutive year, newest first: [(2025, 2026), (2024, 2025), ...]
    year_pairs = list(reversed([(years_sorted[i], years_sorted[i + 1])
                                for i in range(len(years_sorted) - 1)]))

    # ── Title ────────────────────────────────────────────────────────────────
    row = 1
    ws.cell(row, 1, f"{portfolio_name} — Portfolio Summary ({num_props} Properties)")
    ws.cell(row, 1).font = BOLD_FONT
    row += 2

    # ── Transposed KPI table: KPI label | Q1-2024 | Q2-2024 | ... | YoY cols ─
    num_period_cols = len(period_labels)
    total_kpi_cols = 1 + num_period_cols + len(year_pairs) * 2

    # Header row
    kpi_header_row = row
    ws.cell(row, 1, "KPI")
    for col_idx, lbl in enumerate(period_labels, 2):
        cell = ws.cell(row, col_idx, lbl)
        cell.alignment = Alignment(horizontal="center")
    # YoY column headers — two columns per year pair
    for pair_idx, (prev_yr, curr_yr) in enumerate(year_pairs):
        yoy_delta_col = 2 + num_period_cols + pair_idx * 2
        yoy_pct_col   = yoy_delta_col + 1
        cell = ws.cell(row, yoy_delta_col, f"Bud Δ\n{prev_yr}→{curr_yr}")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell = ws.cell(row, yoy_pct_col, f"Bud Δ %\n{prev_yr}→{curr_yr}")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    style_header_row(ws, row, total_kpi_cols)
    row += 1

    _PARENT_FILL = PatternFill("solid", fgColor="2E75B6")
    _PARENT_FONT = Font(bold=True, color="FFFFFF", size=10)
    _NON_GROUPED = {"Income/Unit", "Expense/Unit", "NOI/Unit"}

    # Pre-compute per-period unit counts and excluded properties for per-unit notes
    per_unit_notes: dict[str, tuple] = {}  # period_label -> (total_units, excluded_props_list)
    for (yr, q) in quarters:
        q_kpis_nu = [k for k in _kpis_for_quarter(kpis, yr, q) if not k.is_carveout]
        prop_units: dict[str, int] = {}
        for k in q_kpis_nu:
            if k.total_units is not None:
                prop_units[k.property_name] = k.total_units
        all_q_props = {k.property_name for k in q_kpis_nu}
        total_u = sum(prop_units.values()) if prop_units else None
        excluded_u = sorted(all_q_props - set(prop_units.keys()))
        per_unit_notes[_quarter_label(yr, q)] = (total_u, excluded_u)

    # KPI data rows
    for entry in _DASHBOARD_KPI_ROWS:
        if entry is None:
            row += 1
            continue
        label, key, fmt = entry
        cell = ws.cell(row, 1, label)

        # Leakage Gap: override comment with interpretation guide
        if label == "Leakage Gap":
            c = Comment(_LEAKAGE_GAP_NOTE, "Portfolio Analyzer")
            c.width = 420
            c.height = 160
            cell.comment = c
        else:
            add_kpi_comment(cell, label)

        for col_idx, period_lbl in enumerate(period_labels, 2):
            val = period_aggs[period_lbl].get(key)
            c = ws.cell(row, col_idx, val)
            if fmt:
                c.number_format = fmt

            # Leakage Gap: apply traffic-light fill to each period value
            if label == "Leakage Gap" and val is not None:
                if 0 <= val <= 0.02:
                    c.fill = _LEAKAGE_GREEN
                elif 0.03 < val <= 0.05:
                    c.fill = _LEAKAGE_YELLOW
                elif val > 0.05:
                    c.fill = _LEAKAGE_RED

            # Per-unit rows: add a comment to each value cell with units used
            if label in ("Income/Unit", "Expense/Unit", "NOI/Unit"):
                total_u, excluded_u = per_unit_notes.get(period_lbl, (None, []))
                if total_u is not None:
                    note_lines = [f"Calculated using {total_u:,} total units."]
                    if excluded_u:
                        note_lines.append(
                            "\nProperties excluded (no occupancy data):\n"
                            + "\n".join(f"  • {p}" for p in excluded_u)
                        )
                    cm = Comment("\n".join(note_lines), "Portfolio Analyzer")
                    cm.width = 320
                    cm.height = 100 + 14 * len(excluded_u)
                    c.comment = cm

        # YoY columns — annual aggregate comparisons
        for pair_idx, (prev_yr, curr_yr) in enumerate(year_pairs):
            yoy_delta_col = 2 + num_period_cols + pair_idx * 2
            yoy_pct_col   = yoy_delta_col + 1
            bud_key = _BUDGET_YOY_KEY.get(key)
            if bud_key and (key in _YOY_CURRENCY_KEYS or key in _YOY_PCT_KEYS):
                prev_val = year_aggs[prev_yr].get(bud_key)
                curr_val = year_aggs[curr_yr].get(bud_key)
                delta = (
                    curr_val - prev_val
                    if (curr_val is not None and prev_val is not None)
                    else None
                )
                delta_cell = ws.cell(row, yoy_delta_col, delta)
                delta_cell.number_format = fmt  # $ or % format matching the KPI row
                if delta is not None:
                    fav = key in _YOY_FAVORABLE_IF_POSITIVE
                    apply_variance_fill(delta_cell, delta, favorable_is_positive=fav)
                if key in _YOY_CURRENCY_KEYS:
                    # % change column only for dollar-amount KPIs
                    pct_chg = (
                        delta / abs(prev_val)
                        if (delta is not None and prev_val)
                        else None
                    )
                    pct_cell = ws.cell(row, yoy_pct_col, pct_chg)
                    pct_cell.number_format = PCT_FMT
                    if pct_chg is not None:
                        fav = key in _YOY_FAVORABLE_IF_POSITIVE
                        apply_variance_fill(pct_cell, pct_chg, favorable_is_positive=fav)
                # For PCT keys the % Chg column stays blank (pp delta is sufficient)

        if label in _GROUP_PARENTS:
            for col_idx in range(1, total_kpi_cols + 1):
                gc = ws.cell(row, col_idx)
                gc.fill = _PARENT_FILL
                gc.font = _PARENT_FONT
        elif label not in _NON_GROUPED:
            ws.row_dimensions[row].outline_level = 1
            ws.row_dimensions[row].hidden = True

        row += 1

    kpi_end_row = row - 1

    # ── Portfolio Summary (right of KPI table, merged cells) ─────────────────
    summary_start_col = total_kpi_cols + 2
    summary_end_col = total_kpi_cols + 7

    ws.merge_cells(
        start_row=kpi_header_row, start_column=summary_start_col,
        end_row=kpi_header_row, end_column=summary_end_col,
    )
    hdr_cell = ws.cell(kpi_header_row, summary_start_col, "Portfolio Summary")
    hdr_cell.fill = PatternFill("solid", fgColor="1F4E79")
    hdr_cell.font = Font(bold=True, color="FFFFFF", size=11)

    ws.merge_cells(
        start_row=kpi_header_row + 1, start_column=summary_start_col,
        end_row=kpi_end_row, end_column=summary_end_col,
    )
    summary_text = _generate_portfolio_summary(kpis, eco_occ_target)
    txt_cell = ws.cell(kpi_header_row + 1, summary_start_col, summary_text)
    txt_cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
    txt_cell.font = Font(name="Arial", size=10, color="000000")

    row += 2

    # ── Portfolio AR Aging Summary ───────────────────────────────────────────
    row = _write_dashboard_ar_summary(ws, ar_rows, row, kpi_props=props)
    if ar_rows:
        row += 1

    # ── Top 5 Positive NOI Variance ──────────────────────────────────────────
    ws.cell(row, 1, f"Top Positive NOI Variances ({num_props} Properties Analyzed)").font = BOLD_FONT
    row += 1
    row = _write_top_noi_table(ws, kpis, row, top_n=5, favorable=True)
    row += 2

    # ── Top 5 Negative NOI Variance ──────────────────────────────────────────
    ws.cell(row, 1, f"Top Negative NOI Variances ({num_props} Properties Analyzed)").font = BOLD_FONT
    row += 1
    row = _write_top_noi_table(ws, kpis, row, top_n=5, favorable=False)
    row += 2

    # ── Properties Below Eco Occ Target ─────────────────────────────────────
    if use_budget_eco_occ:
        section_title = "Properties Below Budgeted Economic Occupancy"
    else:
        section_title = f"Properties Below Economic Occupancy Target ({eco_occ_target:.0%})"
    ws.cell(row, 1, section_title).font = BOLD_FONT
    row += 1
    _write_below_target_table(ws, kpis, row, eco_occ_target,
                              use_budget_eco_occ=use_budget_eco_occ)

    # ── Recently Stabilised / New Properties ─────────────────────────────────
    partial_kpis = [k for k in kpis if k.is_partial_year and not k.is_carveout]
    if partial_kpis:
        # row position: find next empty row after the below-target table
        # (use a large offset; the below-target table is small)
        py_row = ws.max_row + 3
        ws.cell(py_row, 1,
            "Recently Stabilised / New Properties "
            "(partial-year data — excluded from all portfolio totals and YoY comparisons)"
        ).font = BOLD_FONT
        py_row += 1

        py_headers = [
            "Property", "PM", "City", "Tenancy Type", "Total Units",
            "Actual Income", "Actual Expenses", "Actual NOI",
            "Budget NOI", "NOI Var vs Budget", "NOI Var %", "Eco Occ %",
        ]
        for ci, h in enumerate(py_headers, 1):
            ws.cell(py_row, ci, h)
        style_header_row(ws, py_row, len(py_headers), fill=_DARK_HDR_FILL, font=_DARK_HDR_FONT)
        py_row += 1

        # Aggregate per partial-year property for the latest quarter
        py_quarters = _get_sorted_quarters(partial_kpis)
        if py_quarters:
            _py_yr, _py_q = py_quarters[0]
            py_q_kpis = _kpis_for_quarter(partial_kpis, _py_yr, _py_q)
            py_groups: dict[str, list] = {}
            for k in py_q_kpis:
                py_groups.setdefault(k.property_name, []).append(k)

            for rank, (prop, pklist) in enumerate(sorted(py_groups.items()), 0):
                agg = _aggregate(pklist)
                row_fill = _ICE_BLUE_FILL if rank % 2 == 0 else _WHITE_FILL
                for ci in range(1, len(py_headers) + 1):
                    ws.cell(py_row, ci).fill = row_fill
                ws.cell(py_row, 1, prop)
                ws.cell(py_row, 2, pklist[0].pm_name)
                ws.cell(py_row, 3, pklist[0].city)
                ws.cell(py_row, 4, pklist[0].tenancy_type)
                ws.cell(py_row, 5, agg.get("total_units") or "N/A")
                _c(ws, py_row, 6,  agg.get("actual_income"),   CURRENCY_FMT)
                _c(ws, py_row, 7,  agg.get("actual_expenses"),  CURRENCY_FMT)
                _c(ws, py_row, 8,  agg.get("actual_noi"),       CURRENCY_FMT)
                _c(ws, py_row, 9,  agg.get("budget_noi"),       CURRENCY_FMT)
                noi_var = (agg.get("actual_noi") - agg.get("budget_noi")
                           if agg.get("actual_noi") is not None and agg.get("budget_noi") is not None
                           else None)
                noi_var_pct = (noi_var / abs(agg.get("budget_noi"))
                               if noi_var is not None and agg.get("budget_noi")
                               else None)
                _c(ws, py_row, 10, noi_var,     CURRENCY_FMT)
                _c(ws, py_row, 11, noi_var_pct, VAR_PCT_FMT)
                _c(ws, py_row, 12, agg.get("eco_occ_pct"), PCT_FMT)
                if noi_var is not None:
                    apply_variance_fill(ws.cell(py_row, 10), noi_var, favorable_is_positive=True)
                py_row += 1

    # ── Column widths ────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 28
    for col_idx in range(2, total_kpi_cols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16
    ws.column_dimensions[get_column_letter(total_kpi_cols + 1)].width = 4
    for col_idx in range(summary_start_col, summary_end_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    # No frozen panes — different sections have different column layouts


# ─── Property Analysis ────────────────────────────────────────────────────────

def _build_property_analysis(wb, kpis, portfolio_name, eco_occ_target):
    ws = wb.create_sheet("Property Quarterly KPIs")
    props = {k.property_name for k in kpis if not k.is_carveout}
    num_props = len(props)

    headers = [
        "Property", "Property Manager", "City", "Tenancy Type", "Period", "Total Units",
        "Actual Income", "Budget Income", "Income Variance", "Income Variance %",
        "Actual Expenses", "Budget Expenses", "Expense Variance", "Expense Variance %",
        "Actual NOI", "Budget NOI", "NOI Variance", "NOI Variance %",
        "Top NOI Driver 1", "Top NOI Driver 2",
        "Eco Occ %", "Budget Eco Occ %", "Eco Occ Variance",
        "GPR", "Vacancy", "Concessions", "Bad Debt",
        "Physical Occ %", "Leakage Gap",
        "Income/Unit", "Expense/Unit", "NOI/Unit",
        "Below Eco Occ Target?",
    ]
    ws.cell(1, 1, f"{portfolio_name} — Property Quarterly KPIs ({num_props} Properties)").font = BOLD_FONT
    for col_idx, hdr in enumerate(headers, 1):
        cell = ws.cell(2, col_idx, hdr)
        add_kpi_comment(cell, hdr)
    style_header_row(ws, 2, len(headers))

    quarters = _get_sorted_quarters(kpis)
    row = 3
    data_row_idx = 0  # used for alternating fill colours
    for (yr, q) in quarters:
        period_lbl = _quarter_label(yr, q)
        q_kpis = _kpis_for_quarter(kpis, yr, q)

        prop_groups: dict[str, list] = {}
        for k in q_kpis:
            prop_groups.setdefault(k.property_name, []).append(k)

        for prop_name in sorted(prop_groups):
            group = prop_groups[prop_name]
            agg = _aggregate(group)
            pm = group[0].pm_name
            is_below = (agg.get("eco_occ_pct") or 0) < eco_occ_target and agg.get("eco_occ_pct") is not None
            row_data = [
                prop_name, pm, group[0].city, group[0].tenancy_type,
                period_lbl, agg.get("total_units") or "Not Available",
                agg.get("actual_income"), agg.get("budget_income"),
                agg.get("income_variance"), agg.get("income_variance_pct"),
                agg.get("actual_expenses"), agg.get("budget_expenses"),
                agg.get("expense_variance"), agg.get("expense_variance_pct"),
                agg.get("actual_noi"), agg.get("budget_noi"),
                agg.get("noi_variance"), agg.get("noi_variance_pct"),
                group[0].top_noi_driver_1, group[0].top_noi_driver_2,
                agg.get("eco_occ_pct"), agg.get("budget_eco_occ_pct"), agg.get("eco_occ_variance"),
                agg.get("gpr"), agg.get("vacancy"), agg.get("concessions"), agg.get("bad_debt"),
                agg.get("physical_occ_pct"), agg.get("leakage_gap"),
                agg.get("income_per_unit"), agg.get("expense_per_unit"), agg.get("noi_per_unit"),
                "YES" if is_below else "no",
            ]
            # Apply alternating ice-blue fill before writing values
            row_fill = _ICE_BLUE_FILL if data_row_idx % 2 == 0 else _WHITE_FILL
            for ci in range(1, len(headers) + 1):
                ws.cell(row, ci).fill = row_fill
            for col_idx, val in enumerate(row_data, 1):
                ws.cell(row, col_idx, val)
            _apply_property_row_formats(ws, row, headers, agg)
            # Re-apply variance fills so they override the ice-blue base fill
            _apply_property_variance_fills(ws, row, headers, agg)
            data_row_idx += 1
            row += 1

    last_data_row = row - 1
    # Wrap in an Excel Table for structured sorting/filtering
    if last_data_row >= 2:
        tab = Table(
            displayName="PropertyQtrKPIsTable",
            ref=f"A2:{get_column_letter(len(headers))}{last_data_row}",
        )
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False, showLastColumn=False,
            showRowStripes=False, showColumnStripes=False,
        )
        ws.add_table(tab)

    ws.freeze_panes = "F3"
    _autofit_columns(ws)


# ─── Property Monthly KPIs ────────────────────────────────────────────────────

def _build_monthly_kpis(wb, kpis):
    ws = wb.create_sheet("Property Monthly KPIs")
    headers = [
        "Property", "Property Manager", "City", "Tenancy Type", "Year", "Month", "Period",
        "Total Units",
        "Actual Income", "Budget Income", "Income Variance",
        "Actual Expenses", "Budget Expenses", "Expense Variance",
        "Actual NOI", "Budget NOI", "NOI Variance", "NOI Variance %",
        "GPR", "Vacancy", "Concessions", "Bad Debt",
        "Net Collectible", "Eco Occ %",
        "Physical Occ %", "Leakage Gap",
        "YoY Physical Occ Var", "YoY Eco Occ Var", "YoY Leakage Gap Change",
        "Income/Unit", "Expense/Unit", "NOI/Unit",
        "Source Key",
    ]
    for col_idx, hdr in enumerate(headers, 1):
        cell = ws.cell(1, col_idx, hdr)
        add_kpi_comment(cell, hdr)
    style_header_row(ws, 1, len(headers))

    row = 2
    for data_row_idx, k in enumerate(sorted(kpis, key=lambda x: (x.property_name, -x.year, -x.month))):
        # Alternating ice-blue fill
        row_fill = _ICE_BLUE_FILL if data_row_idx % 2 == 0 else _WHITE_FILL
        for ci in range(1, len(headers) + 1):
            ws.cell(row, ci).fill = row_fill

        ws.cell(row, 1,  k.property_name)
        ws.cell(row, 2,  k.pm_name)
        ws.cell(row, 3,  k.city)
        ws.cell(row, 4,  k.tenancy_type)
        ws.cell(row, 5,  k.year)
        ws.cell(row, 6,  k.month)
        ws.cell(row, 7,  k.period)
        ws.cell(row, 8,  k.total_units if k.total_units is not None else "N/A")
        _c(ws, row, 9,  k.actual_income,   CURRENCY_FMT)
        _c(ws, row, 10, k.budget_income,   CURRENCY_FMT)
        _c(ws, row, 11, k.income_variance, CURRENCY_FMT)
        _c(ws, row, 12, k.actual_expenses,  CURRENCY_FMT)
        _c(ws, row, 13, k.budget_expenses,  CURRENCY_FMT)
        _c(ws, row, 14, k.expense_variance, CURRENCY_FMT)
        _c(ws, row, 15, k.actual_noi,       CURRENCY_FMT)
        _c(ws, row, 16, k.budget_noi,       CURRENCY_FMT)
        _c(ws, row, 17, k.noi_variance,     CURRENCY_FMT)
        _c(ws, row, 18, k.noi_variance_pct, VAR_PCT_FMT)
        _c(ws, row, 19, k.gpr,              CURRENCY_FMT)
        _c(ws, row, 20, k.vacancy,          CURRENCY_FMT)
        _c(ws, row, 21, k.concessions,      CURRENCY_FMT)
        _c(ws, row, 22, k.bad_debt,         CURRENCY_FMT)
        _c(ws, row, 23, k.net_collectible,  CURRENCY_FMT)
        _c(ws, row, 24, k.eco_occ_pct,      PCT_FMT)
        _c(ws, row, 25, k.physical_occ_pct, PCT_FMT)
        _c(ws, row, 26, k.leakage_gap,      PCT_FMT)
        _c(ws, row, 27, k.yoy_physical_occ_variance, PCT_FMT)
        _c(ws, row, 28, k.yoy_eco_occ_variance,      PCT_FMT)
        _c(ws, row, 29, k.yoy_leakage_gap_change,    PCT_FMT)
        _c(ws, row, 30, k.income_per_unit,  CURRENCY_FMT)
        _c(ws, row, 31, k.expense_per_unit, CURRENCY_FMT)
        _c(ws, row, 32, k.noi_per_unit,     CURRENCY_FMT)
        ws.cell(row, 33, k.source_key)
        # Re-apply variance fill to % column only — $ amount variance columns are intentionally unstyled.
        # Only highlight when variance exceeds ±5% threshold.
        apply_variance_fill(ws.cell(row, 18), k.noi_variance_pct,
                            favorable_is_positive=True, threshold=0.05)
        row += 1

    last_data_row = row - 1
    # Wrap in an Excel Table
    if last_data_row >= 1:
        tab = Table(
            displayName="MonthlyKPIsTable",
            ref=f"A1:{get_column_letter(len(headers))}{last_data_row}",
        )
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False, showLastColumn=False,
            showRowStripes=False, showColumnStripes=False,
        )
        ws.add_table(tab)

    ws.freeze_panes = "H2"
    _autofit_columns(ws)


def _c(ws, row, col, value, fmt):
    """Write a cell value and apply a number format."""
    ws.cell(row, col, value).number_format = fmt


# ─── Aggregation ─────────────────────────────────────────────────────────────

def _aggregate(kpis: list[PropertyPeriodKPIs]) -> dict:
    """
    Aggregate numeric KPI fields across a list of PropertyPeriodKPIs records.

    Financial flows (income, expenses, GPR, etc.) are SUMMED — monthly amounts
    accumulate to a period total.

    Rate metrics (physical_occ_pct, eco_occ_pct) use a WEIGHTED approach:
      physical_occ = Σ(occupied_units) / Σ(total_units) across matched months.
    This avoids the >100% error that results from summing occupied_units and
    dividing by a single month's total_units.
    """
    def _sum(field):
        vals = [getattr(k, field) for k in kpis if getattr(k, field) is not None]
        return sum(vals) if vals else None

    actual_income   = _sum("actual_income")
    budget_income   = _sum("budget_income")
    actual_expenses = _sum("actual_expenses")
    budget_expenses = _sum("budget_expenses")

    actual_noi = (actual_income - actual_expenses) if (actual_income is not None and actual_expenses is not None) else None
    budget_noi = (budget_income - budget_expenses) if (budget_income is not None and budget_expenses is not None) else None

    gpr         = _sum("gpr")
    vacancy     = _sum("vacancy")
    concessions = _sum("concessions")
    bad_debt    = _sum("bad_debt")
    net_coll    = (gpr - (vacancy or 0) - (concessions or 0) - (bad_debt or 0)) if gpr is not None else None
    eco_occ     = (net_coll / gpr) if (net_coll is not None and gpr) else None

    # Budget eco occ: average the per-month budget rates across the period
    bud_eco_vals = [k.budget_eco_occ_pct for k in kpis if k.budget_eco_occ_pct is not None]
    bud_eco      = sum(bud_eco_vals) / len(bud_eco_vals) if bud_eco_vals else None
    eco_occ_var  = (eco_occ - bud_eco) if (eco_occ is not None and bud_eco is not None) else None

    # Physical occ: Σ(occupied) / Σ(total) — correct time-weighted average.
    # Using max(total_units) would produce values > 100% when aggregating multiple months.
    _paired = [
        (k.occupied_units, k.total_units)
        for k in kpis
        if k.occupied_units is not None and k.total_units is not None
    ]
    if _paired:
        _occ_sum   = sum(p[0] for p in _paired)
        _total_sum = sum(p[1] for p in _paired)
        phys_occ   = _occ_sum / _total_sum if _total_sum > 0 else None
    else:
        phys_occ = None
    # For per-unit metrics use total unit-months as the denominator:
    #   single property Q1  → units × 3 months  → monthly rate
    #   portfolio     Q1    → Σ(units_per_prop × 3)  → monthly rate
    # This is consistent regardless of the aggregation level.
    total_units = next((k.total_units for k in kpis if k.total_units is not None), None)
    total_unit_months = sum(k.total_units for k in kpis if k.total_units is not None) or None

    def _safe_div(a, b):
        return (a - b) if (a is not None and b is not None) else None

    def _safe_pct(a, b):
        return (a / abs(b)) if (a is not None and b is not None and b != 0) else None

    income_var  = _safe_div(actual_income, budget_income)
    expense_var = _safe_div(actual_expenses, budget_expenses)
    noi_var     = _safe_div(actual_noi, budget_noi)

    income_pu  = (actual_income   / total_unit_months) if (actual_income   is not None and total_unit_months) else None
    expense_pu = (actual_expenses / total_unit_months) if (actual_expenses is not None and total_unit_months) else None
    noi_pu     = (actual_noi      / total_unit_months) if (actual_noi      is not None and total_unit_months) else None

    return dict(
        actual_income=actual_income, budget_income=budget_income,
        income_variance=income_var, income_variance_pct=_safe_pct(income_var, budget_income),
        actual_expenses=actual_expenses, budget_expenses=budget_expenses,
        expense_variance=expense_var, expense_variance_pct=_safe_pct(expense_var, budget_expenses),
        actual_noi=actual_noi, budget_noi=budget_noi,
        noi_variance=noi_var, noi_variance_pct=_safe_pct(noi_var, budget_noi),
        gpr=gpr, vacancy=vacancy, concessions=concessions, bad_debt=bad_debt,
        net_collectible=net_coll, eco_occ_pct=eco_occ, budget_eco_occ_pct=bud_eco,
        eco_occ_variance=eco_occ_var,
        total_units=total_units, physical_occ_pct=phys_occ,
        leakage_gap=(phys_occ - eco_occ) if (phys_occ is not None and eco_occ is not None) else None,
        income_per_unit=income_pu, expense_per_unit=expense_pu, noi_per_unit=noi_pu,
    )


# ─── Dashboard section helpers ────────────────────────────────────────────────

def _write_top_noi_table(ws, kpis, start_row, top_n, favorable):
    """Rank properties by Actual NOI vs Budget NOI for the latest quarter.

    Columns: Rank | Property | PM | City | Actual NOI | Budget NOI |
             NOI Var vs Budget $ | NOI Var vs Budget % |
             Income Var vs Budget | Expense Var vs Budget
    """
    headers = [
        "Rank", "Property", "PM", "City",
        "Actual NOI", "Budget NOI", "NOI Var vs Budget", "NOI Var %",
        "Income Var vs Budget", "Expense Var vs Budget",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(start_row, col, h)
    style_header_row(ws, start_row, len(headers), fill=_DARK_HDR_FILL, font=_DARK_HDR_FONT)
    ws.row_dimensions[start_row].height = 26
    row = start_row + 1

    # Use the latest quarter's data
    quarters = _get_sorted_quarters(kpis)   # newest first
    if not quarters:
        ws.cell(row, 1, "No data available.")
        return row + 1

    latest_yr, latest_q = quarters[0]
    q_kpis = [k for k in _kpis_for_quarter(kpis, latest_yr, latest_q)
              if not k.is_carveout]

    # Aggregate per property for the latest quarter
    prop_groups: dict[str, list] = {}
    for k in q_kpis:
        prop_groups.setdefault(k.property_name, []).append(k)

    variances = []
    for prop, pklist in prop_groups.items():
        agg = _aggregate(pklist)
        actual_noi  = agg.get("actual_noi")
        budget_noi  = agg.get("budget_noi")
        if actual_noi is None or budget_noi is None:
            continue
        noi_var     = actual_noi - budget_noi
        noi_var_pct = noi_var / abs(budget_noi) if budget_noi else None
        income_var  = agg.get("income_variance")    # actual_income - budget_income
        expense_var = agg.get("expense_variance")   # actual_expenses - budget_expenses
        variances.append((
            noi_var, prop, pklist[0].pm_name, pklist[0].city,
            actual_noi, budget_noi, noi_var_pct, income_var, expense_var,
        ))

    variances.sort(key=lambda x: x[0], reverse=favorable)

    for rank, (noi_var, prop, pm, city,
               actual_noi, budget_noi, noi_var_pct,
               income_var, expense_var) in enumerate(variances[:top_n], 1):
        row_fill = _ICE_BLUE_FILL if rank % 2 == 1 else _WHITE_FILL
        for ci in range(1, len(headers) + 1):
            ws.cell(row, ci).fill = row_fill
        ws.cell(row, 1, rank)
        ws.cell(row, 2, prop)
        ws.cell(row, 3, pm)
        ws.cell(row, 4, city)
        _c(ws, row, 5, actual_noi,  CURRENCY_FMT)
        _c(ws, row, 6, budget_noi,  CURRENCY_FMT)
        _c(ws, row, 7, noi_var,     CURRENCY_FMT)
        _c(ws, row, 8, noi_var_pct, VAR_PCT_FMT)
        # Income variance: positive = above budget (favorable)
        _c(ws, row, 9,  income_var,  CURRENCY_FMT)
        if income_var is not None:
            apply_variance_fill(ws.cell(row, 9),  income_var,  favorable_is_positive=True)
        # Expense variance: positive = over budget (unfavorable)
        _c(ws, row, 10, expense_var, CURRENCY_FMT)
        if expense_var is not None:
            apply_variance_fill(ws.cell(row, 10), expense_var, favorable_is_positive=False)
        # NOI variance fill
        if noi_var is not None:
            apply_variance_fill(ws.cell(row, 7), noi_var, favorable_is_positive=True)
        row += 1

    if row == start_row + 1:
        ws.cell(row, 1, "No properties with both actual and budget NOI for the latest period.")
        row += 1
    return row


def _eco_occ_drivers(agg: dict) -> tuple[str, str]:
    components = [
        (agg.get("vacancy") or 0,     "Vacancy"),
        (agg.get("concessions") or 0, "Concessions"),
        (agg.get("bad_debt") or 0,    "Bad Debt"),
    ]
    # Sort by absolute dollar amount descending
    components.sort(key=lambda x: abs(x[0]), reverse=True)
    def fmt(amt, name):
        return f"{name} (${abs(amt):,.0f})" if amt else name
    d1 = fmt(components[0][0], components[0][1]) if components else ""
    d2 = fmt(components[1][0], components[1][1]) if len(components) > 1 else ""
    return d1, d2


def _write_below_target_table(ws, kpis, start_row, eco_occ_target,
                              use_budget_eco_occ=False):
    quarters = _get_sorted_quarters(kpis)
    if not quarters:
        ws.cell(start_row + 1, 1, "No data available.")
        return start_row + 2

    # With quarters sorted newest-first, index 0 is the most recent quarter
    latest_q_yr, latest_q = quarters[0]
    q_kpis = _kpis_for_quarter(kpis, latest_q_yr, latest_q)

    # Aggregate per property for this quarter
    prop_groups: dict = {}
    for k in q_kpis:
        prop_groups.setdefault(k.property_name, []).append(k)

    below = []
    for prop, pklist in prop_groups.items():
        agg = _aggregate(pklist)
        pm = pklist[0].pm_name
        city = pklist[0].city
        tenancy = pklist[0].tenancy_type
        eco_occ = agg.get("eco_occ_pct")
        if eco_occ is None:
            continue
        bud = agg.get("budget_eco_occ_pct")
        effective_target = (bud if (use_budget_eco_occ and bud is not None)
                            else eco_occ_target)
        if eco_occ < effective_target:
            below.append((eco_occ, prop, pm, city, tenancy, agg, effective_target))

    below.sort(key=lambda x: x[0])  # worst first

    if use_budget_eco_occ:
        headers = ["Property", "PM", "City", "Tenancy Type", "Eco Occ %",
                   "Budget Eco Occ %", "Variance to Budget", "Driver 1", "Driver 2"]
    else:
        headers = ["Property", "PM", "City", "Tenancy Type", "Eco Occ %",
                   "Target", "Variance to Target", "Driver 1", "Driver 2"]
    for col, h in enumerate(headers, 1):
        ws.cell(start_row, col, h)
    style_header_row(ws, start_row, len(headers), fill=_DARK_HDR_FILL, font=_DARK_HDR_FONT)
    ws.row_dimensions[start_row].height = 26
    row = start_row + 1

    for rank, (eco_occ, prop, pm, city, tenancy, agg, effective_target) in enumerate(below):
        # Alternating ice-blue rows
        row_fill = _ICE_BLUE_FILL if rank % 2 == 0 else _WHITE_FILL
        for ci in range(1, len(headers) + 1):
            ws.cell(row, ci).fill = row_fill
        d1, d2 = _eco_occ_drivers(agg)
        ws.cell(row, 1, prop)
        ws.cell(row, 2, pm)
        ws.cell(row, 3, city)
        ws.cell(row, 4, tenancy)
        _c(ws, row, 5, eco_occ, PCT_FMT)
        _c(ws, row, 6, effective_target, PCT_FMT)
        _c(ws, row, 7, eco_occ - effective_target, PCT_FMT)
        ws.cell(row, 8, d1)
        ws.cell(row, 9, d2)
        row += 1

    if row == start_row + 1:
        if use_budget_eco_occ:
            ws.cell(row, 1, "All properties at or above their budgeted economic occupancy")
        else:
            ws.cell(row, 1, f"All properties at or above {eco_occ_target:.0%} target")
        row += 1

    # Autofilter on the header row (covers the data range below it)
    last_data_row = row - 1
    ws.auto_filter.ref = (
        f"A{start_row}:{get_column_letter(len(headers))}{last_data_row}"
    )
    return row


def _generate_portfolio_summary(kpis: list, eco_occ_target: float) -> str:
    """Generate a comprehensive 5-6 sentence portfolio summary."""
    years = sorted({k.year for k in kpis})
    if len(years) < 2:
        yr = years[0] if years else "N/A"
        total_noi = sum(k.actual_noi or 0 for k in kpis)
        return (f"Portfolio NOI for {yr} totaled ${total_noi:,.0f}. "
                "Prior year data not available for trend comparison.")

    curr_yr, prev_yr = years[-1], years[-2]

    # YoY aggregates
    curr_noi = sum(k.actual_noi or 0 for k in kpis if k.year == curr_yr)
    prev_noi = sum(k.actual_noi or 0 for k in kpis if k.year == prev_yr)
    curr_inc = sum(k.actual_income or 0 for k in kpis if k.year == curr_yr)
    prev_inc = sum(k.actual_income or 0 for k in kpis if k.year == prev_yr)
    curr_exp = sum(k.actual_expenses or 0 for k in kpis if k.year == curr_yr)
    prev_exp = sum(k.actual_expenses or 0 for k in kpis if k.year == prev_yr)

    noi_var = curr_noi - prev_noi
    noi_dir = "improved" if noi_var >= 0 else "declined"
    noi_pct = noi_var / abs(prev_noi) * 100 if prev_noi else 0
    inc_var = curr_inc - prev_inc
    exp_var = curr_exp - prev_exp
    primary_driver = "income" if abs(inc_var) >= abs(exp_var) else "expense"
    driver_dir = "favorable" if (primary_driver == "income" and inc_var >= 0) or (primary_driver == "expense" and exp_var <= 0) else "unfavorable"

    # Per-property YoY NOI variances
    curr_by_prop: dict = {}
    prev_by_prop: dict = {}
    for k in kpis:
        if k.actual_noi is None:
            continue
        if k.year == curr_yr:
            curr_by_prop[k.property_name] = curr_by_prop.get(k.property_name, 0) + k.actual_noi
        elif k.year == prev_yr:
            prev_by_prop[k.property_name] = prev_by_prop.get(k.property_name, 0) + k.actual_noi

    prop_vars = sorted(
        [(curr_by_prop[p] - prev_by_prop[p], p)
         for p in curr_by_prop if p in prev_by_prop],
        key=lambda x: x[0], reverse=True
    )
    top5_pos = prop_vars[:5]
    top5_neg = list(reversed(prop_vars[-5:])) if len(prop_vars) >= 5 else list(reversed(prop_vars))
    top5_neg = [x for x in top5_neg if x[0] < 0]

    # Latest quarter: properties below eco occ target + leakage
    quarters = _get_sorted_quarters(kpis)
    n_below = 0
    top5_leakage = []
    latest_q_label = ""
    if quarters:
        latest_q_yr, latest_q = quarters[-1]
        latest_q_label = f"Q{latest_q} {latest_q_yr}"
        q_kpis = _kpis_for_quarter(kpis, latest_q_yr, latest_q)
        prop_groups: dict = {}
        for k in q_kpis:
            prop_groups.setdefault(k.property_name, []).append(k)
        leakage_data = []
        for prop, pklist in prop_groups.items():
            agg = _aggregate(pklist)
            if agg.get("eco_occ_pct") is not None and agg["eco_occ_pct"] < eco_occ_target:
                n_below += 1
            if agg.get("leakage_gap") is not None and agg["leakage_gap"] > 0:
                leakage_data.append((agg["leakage_gap"], prop))
        leakage_data.sort(reverse=True)
        top5_leakage = leakage_data[:5]

    # Build sentences
    s1 = (f"Portfolio NOI {noi_dir} by ${abs(noi_var):,.0f} ({abs(noi_pct):.1f}%) from {prev_yr} to {curr_yr}, "
          f"primarily driven by {driver_dir} {primary_driver} trends "
          f"(income {'+' if inc_var >= 0 else ''}{inc_var:,.0f}, expenses {'+' if exp_var >= 0 else ''}{exp_var:,.0f} YoY).")

    if top5_pos:
        pos_list = ", ".join(f"{p} (${v:+,.0f})" for v, p in top5_pos)
        s2 = f"Top 5 positive NOI variances: {pos_list}."
    else:
        s2 = "No properties showed positive NOI variance year-over-year."

    if top5_neg:
        neg_list = ", ".join(f"{p} (${v:+,.0f})" for v, p in top5_neg)
        s3 = f"Top 5 negative NOI variances: {neg_list}."
    else:
        s3 = "No properties showed negative NOI variance year-over-year."

    s4 = (f"As of {latest_q_label}, {n_below} propert{'y' if n_below == 1 else 'ies'} "
          f"{'was' if n_below == 1 else 'were'} below the {eco_occ_target:.0%} economic occupancy target, "
          f"indicating revenue collection gaps from vacancy loss, concessions, or bad debt.")

    if top5_leakage:
        leak_list = ", ".join(f"{p} ({gap:.1%})" for gap, p in top5_leakage)
        s5 = (f"Top 5 leakage gap properties — units are occupied but rent is not being fully "
              f"collected — were: {leak_list}.")
    else:
        s5 = "No properties showed a positive leakage gap in the current quarter."

    total_props = len({k.property_name for k in kpis})
    below_pct = n_below / total_props * 100 if total_props else 0
    if noi_var < 0 and below_pct > 30:
        health = "under stress, with both declining NOI and widespread economic occupancy challenges requiring immediate management attention"
    elif noi_var >= 0 and below_pct < 20:
        health = "in generally healthy condition, with improving NOI and manageable occupancy levels across the portfolio"
    else:
        health = "showing mixed performance, with select properties driving outsized variance and warranting targeted review"

    s6 = f"Overall, the portfolio is {health}."

    return "\n\n".join([s1, s2, s3, s4, s5, s6])


def _write_dashboard_ar_summary(ws, ar_rows: list, start_row: int,
                                kpi_props: set | None = None) -> int:
    """Write portfolio-level AR Aging summary block on the Dashboard tab."""
    if not ar_rows:
        return start_row

    periods = sorted({(r.year, r.month) for r in ar_rows})
    periods_set = set(periods)
    col_seq = []
    for (yr, mo) in periods:
        col_seq.append(("period", yr, mo))
        if (yr - 1, mo) in periods_set:
            col_seq.append(("yoy", yr, mo))

    # Properties present in the main KPI analysis but absent from AR data
    ar_all_props = {r.property_name for r in ar_rows}
    excluded_from_ar = sorted((kpi_props or set()) - ar_all_props)

    row = start_row
    for rtype in ("Tenant Rent", "Subsidy"):
        rtype_rows = [r for r in ar_rows if r.receivable_type == rtype]
        if not rtype_rows:
            continue

        prop_count = len({r.property_name for r in rtype_rows})
        header_text = (
            f"Portfolio AR Summary — {rtype} "
            f"({prop_count} Propert{'y' if prop_count == 1 else 'ies'})"
        )
        title_cell = ws.cell(row, 1, header_text)
        title_cell.font = BOLD_FONT
        # Add a comment listing properties not included in AR data
        if excluded_from_ar:
            note = (
                "Properties in analysis but excluded from this AR section\n"
                "(no AR aging file uploaded for these properties):\n"
                + "\n".join(f"  • {p}" for p in excluded_from_ar)
            )
            cm = Comment(note, "Portfolio Analyzer")
            cm.width = 320
            cm.height = 80 + 14 * len(excluded_from_ar)
            title_cell.comment = cm
        row += 1

        num_cols = 1 + len(col_seq)
        ws.cell(row, 1, "Metric")
        for ci, (ctype, yr, mo) in enumerate(col_seq, 2):
            lbl = (f"YoY Δ {_ar_period_label(yr, mo)}" if ctype == "yoy"
                   else _ar_period_label(yr, mo))
            ws.cell(row, ci, lbl)
        style_header_row(ws, row, num_cols)
        row += 1

        def _agg(yr, mo):
            rows = [r for r in rtype_rows if r.year == yr and r.month == mo]
            if not rows:
                return None
            charge = sum(r.charge_amount for r in rows)
            over_60 = sum(r.owed_61_90 + r.owed_over_90 for r in rows)
            return {
                "current_owed": sum(r.current_owed for r in rows),
                "prepayments":  sum(r.prepayments for r in rows),
                "pct_overdue":  (over_60 / charge) if charge > 0 else 0.0,
            }

        period_aggs_ar = {(yr, mo): _agg(yr, mo) for (yr, mo) in periods}

        metric_defs = [
            # (label, key, fmt, fav_pos, apply_yoy_fill)
            # $ amount rows carry no fill; only % >60 Days YoY delta gets shading
            ("Current Owed", "current_owed", CURRENCY_FMT, False, False),
            ("Pre-payments",  "prepayments",  CURRENCY_FMT, False, False),
            ("% >60 Days",    "pct_overdue",  PCT_FMT,      False, True),
        ]
        for m_idx, (label, key, fmt, fav_pos, apply_yoy_fill) in enumerate(metric_defs):
            # Alternating ice-blue fill across data rows
            row_fill = _ICE_BLUE_FILL if m_idx % 2 == 0 else _WHITE_FILL
            for ci in range(1, num_cols + 1):
                ws.cell(row, ci).fill = row_fill

            ws.cell(row, 1, label).font = BOLD_FONT
            for ci, (ctype, yr, mo) in enumerate(col_seq, 2):
                if ctype == "period":
                    a = period_aggs_ar.get((yr, mo))
                    val = a[key] if a else None
                    _c(ws, row, ci, val, fmt)
                else:
                    curr = period_aggs_ar.get((yr, mo))
                    prev = period_aggs_ar.get((yr - 1, mo))
                    if curr and prev and curr.get(key) is not None and prev.get(key) is not None:
                        delta = curr[key] - prev[key]
                        _c(ws, row, ci, delta, fmt)
                        if apply_yoy_fill:
                            # >5% threshold: only colour % >60 Days delta when it exceeds ±5pp
                            apply_variance_fill(ws.cell(row, ci), delta,
                                                favorable_is_positive=fav_pos, threshold=0.05)
            row += 1

        row += 2  # blank rows between blocks

    return row


def _apply_property_row_formats(ws, row, headers, agg):
    fmt_map = {
        "Actual Income": CURRENCY_FMT, "Budget Income": CURRENCY_FMT,
        "Income Variance": CURRENCY_FMT, "Income Variance %": VAR_PCT_FMT,
        "Actual Expenses": CURRENCY_FMT, "Budget Expenses": CURRENCY_FMT,
        "Expense Variance": CURRENCY_FMT, "Expense Variance %": VAR_PCT_FMT,
        "Actual NOI": CURRENCY_FMT, "Budget NOI": CURRENCY_FMT,
        "NOI Variance": CURRENCY_FMT, "NOI Variance %": VAR_PCT_FMT,
        "Eco Occ %": PCT_FMT, "Budget Eco Occ %": PCT_FMT, "Eco Occ Variance": PCT_FMT,
        "Physical Occ %": PCT_FMT, "Leakage Gap": PCT_FMT,
        "GPR": CURRENCY_FMT, "Vacancy": CURRENCY_FMT,
        "Concessions": CURRENCY_FMT, "Bad Debt": CURRENCY_FMT,
        "Income/Unit": CURRENCY_FMT, "Expense/Unit": CURRENCY_FMT, "NOI/Unit": CURRENCY_FMT,
    }
    for col_idx, hdr in enumerate(headers, 1):
        fmt = fmt_map.get(hdr)
        if fmt:
            ws.cell(row, col_idx).number_format = fmt


def _apply_property_variance_fills(ws, row, headers, agg):
    """Re-apply green/red variance fills to % variance columns after ice-blue base fill.
    $ amount variance columns (Income Variance, Expense Variance, NOI Variance) are
    intentionally excluded — only % rate columns receive colour shading.
    A ±5% threshold is applied: cells within the neutral band receive no fill."""
    variance_cols = {
        # $ amount variances intentionally omitted (no shading)
        "Income Variance %":  True,   # favorable_is_positive
        "Expense Variance %": False,  # favorable_is_negative (under budget = good)
        "NOI Variance %":     True,
        "Eco Occ Variance":   True,
    }
    key_map = {
        "Income Variance %":  "income_variance_pct",
        "Expense Variance %": "expense_variance_pct",
        "NOI Variance %":     "noi_variance_pct",
        "Eco Occ Variance":   "eco_occ_variance",
    }
    for col_idx, hdr in enumerate(headers, 1):
        if hdr in variance_cols:
            val = agg.get(key_map[hdr])
            apply_variance_fill(ws.cell(row, col_idx), val,
                                favorable_is_positive=variance_cols[hdr],
                                threshold=0.05)


def _build_ar_aging(wb, ar_rows: list, portfolio_name: str, kpis=None) -> None:
    """4-block AR Aging tab: portfolio summaries (TR + Sub) + property analysis (TR + Sub)."""
    ws = wb.create_sheet("AR Aging")

    if not ar_rows:
        ws.cell(1, 1, "No AR Aging data was uploaded for this analysis.")
        ws.cell(1, 1).font = BOLD_FONT
        return

    # Unique periods — newest first so columns read newest-to-oldest
    periods = sorted({(r.year, r.month) for r in ar_rows}, reverse=True)
    periods_set = set(periods)

    # Build column sequence: each period followed by its YoY delta (if prior year present)
    col_seq: list[tuple] = []
    for (yr, mo) in periods:
        col_seq.append(("period", yr, mo))
        if (yr - 1, mo) in periods_set:
            col_seq.append(("yoy", yr, mo))

    # With descending sort, index 0 is the most recent period
    latest_yr, latest_mo = periods[0]
    latest_label = _ar_period_label(latest_yr, latest_mo)
    prior_period = (latest_yr - 1, latest_mo) if (latest_yr - 1, latest_mo) in periods_set else None

    def _agg(rtype: str, yr: int, mo: int) -> dict | None:
        """Sum AR metrics for a receivable type + period. >60 days = owed_61_90 + owed_over_90."""
        rows = [r for r in ar_rows if r.receivable_type == rtype and r.year == yr and r.month == mo]
        if not rows:
            return None
        charge = sum(r.charge_amount for r in rows)
        over_60 = sum(r.owed_61_90 + r.owed_over_90 for r in rows)
        return {
            "current_owed": sum(r.current_owed for r in rows),
            "prepayments":  sum(r.prepayments for r in rows),
            "pct_overdue":  (over_60 / charge) if charge > 0 else None,
        }

    # Bad-debt lookup from financial KPIs: (property_name, year, month) -> $
    _bd: dict[tuple, float] = {}
    if kpis:
        for k in kpis:
            key = (k.property_name, k.year, k.month)
            _bd[key] = (_bd.get(key) or 0.0) + (k.bad_debt or 0.0)

    def _bd_for_period(prop_names, yr: int, mo: int) -> float | None:
        """Sum bad-debt write-offs across a set of properties for a given period."""
        total = sum(_bd.get((p, yr, mo), 0.0) for p in prop_names)
        return total if total else None

    # Unique table name counter to ensure uniqueness across all 4 blocks
    _table_counter = [0]

    def _next_table_name(prefix: str) -> str:
        _table_counter[0] += 1
        return f"{prefix}{_table_counter[0]}"

    row = 1

    # ── Blocks 1 & 2: Portfolio AR Summary ───────────────────────────────────
    for rtype in ("Tenant Rent", "Subsidy"):
        rtype_rows = [r for r in ar_rows if r.receivable_type == rtype]
        if not rtype_rows:
            continue

        prop_count = len({r.property_name for r in rtype_rows})
        header_text = (
            f"{portfolio_name} — Portfolio AR Summary — {rtype} "
            f"({prop_count} Propert{'y' if prop_count == 1 else 'ies'})"
        )
        ws.cell(row, 1, header_text).font = BOLD_FONT
        row += 1

        num_cols = 1 + len(col_seq)
        hdr_row = row
        ws.cell(row, 1, "Metric")
        for ci, (ctype, yr, mo) in enumerate(col_seq, 2):
            lbl = (f"YoY Δ {_ar_period_label(yr, mo)}" if ctype == "yoy"
                   else _ar_period_label(yr, mo))
            ws.cell(row, ci, lbl)
        style_header_row(ws, row, num_cols)
        row += 1
        data_start = row

        # Pre-compute period aggregates for this receivable type
        period_aggs = {}
        for (yr, mo) in periods:
            agg_val = _agg(rtype, yr, mo)
            if agg_val:
                agg_val["bad_debt"] = _bd_for_period(rtype_props_by_period.get((yr, mo), set()), yr, mo)
            period_aggs[(yr, mo)] = agg_val

        # Collect property names for this receivable type / each period
        rtype_props_by_period = {
            (yr, mo): {r.property_name for r in rtype_rows if r.year == yr and r.month == mo}
            for (yr, mo) in periods
        }

        # Four metric rows: Current Owed, Pre-payments, % >60 Days, Bad Debt Write-off
        metric_defs = [
            # (label, key, fmt, fav_pos, apply_yoy_fill)
            # $ amount rows carry no fill; only % >60 Days YoY delta gets shading
            ("Current Owed",      "current_owed", CURRENCY_FMT, False, False),
            ("Pre-payments",      "prepayments",  CURRENCY_FMT, False, False),
            ("% >60 Days",        "pct_overdue",  PCT_FMT,      False, True),
            ("Bad Debt Write-off","bad_debt",      CURRENCY_FMT, False, False),
        ]
        for m_idx, (label, key, fmt, fav_pos, apply_yoy_fill) in enumerate(metric_defs):
            # Alternating ice-blue fill
            row_fill = _ICE_BLUE_FILL if m_idx % 2 == 0 else _WHITE_FILL
            for ci in range(1, num_cols + 1):
                ws.cell(row, ci).fill = row_fill
            ws.cell(row, 1, label).font = BOLD_FONT
            for ci, (ctype, yr, mo) in enumerate(col_seq, 2):
                if ctype == "period":
                    agg_val = period_aggs.get((yr, mo))
                    val = agg_val[key] if agg_val else None
                    _c(ws, row, ci, val, fmt)
                else:
                    curr = period_aggs.get((yr, mo))
                    prev = period_aggs.get((yr - 1, mo))
                    if curr and prev and curr.get(key) is not None and prev.get(key) is not None:
                        delta = curr[key] - prev[key]
                        _c(ws, row, ci, delta, fmt)
                        if apply_yoy_fill:
                            # >5% threshold: only colour % >60 Days delta when it exceeds ±5pp
                            apply_variance_fill(ws.cell(row, ci), delta,
                                                favorable_is_positive=fav_pos, threshold=0.05)
                    else:
                        ws.cell(row, ci, None)
            row += 1

        # Excel Table for portfolio summary block
        last_metric_row = row - 1
        t_name = _next_table_name("PortfolioAR_")
        tab = Table(
            displayName=t_name,
            ref=f"A{hdr_row}:{get_column_letter(num_cols)}{last_metric_row}",
        )
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False, showLastColumn=False,
            showRowStripes=False, showColumnStripes=False,
        )
        ws.add_table(tab)

        row += 2  # blank rows between blocks

    # ── Blocks 3 & 4: Property AR Analysis ───────────────────────────────────
    for rtype in ("Tenant Rent", "Subsidy"):
        rtype_rows = [r for r in ar_rows if r.receivable_type == rtype]
        if not rtype_rows:
            continue

        ws.cell(row, 1, f"Property AR Analysis — {rtype} (As of {latest_label})").font = BOLD_FONT
        row += 1

        prop_headers = [
            "Property", "PM",
            "Current Owed", "Pre-payments", "% >60 Days",
            "YoY $ Δ (Current Owed)", "YoY Δ (% >60)",
            "Bad Debt Write-off",
        ]
        hdr_row = row
        for ci, h in enumerate(prop_headers, 1):
            ws.cell(row, ci, h)
        style_header_row(ws, row, len(prop_headers))
        row += 1
        data_start_row = row

        # Aggregate per property for latest period (sum if multiple files)
        def _prop_agg(prop: str, yr: int, mo: int) -> dict | None:
            rows = [r for r in rtype_rows if r.property_name == prop and r.year == yr and r.month == mo]
            if not rows:
                return None
            charge = sum(r.charge_amount for r in rows)
            over_60 = sum(r.owed_61_90 + r.owed_over_90 for r in rows)
            return {
                "current_owed": sum(r.current_owed for r in rows),
                "prepayments":  sum(r.prepayments for r in rows),
                "pct_overdue":  (over_60 / charge) if charge > 0 else 0.0,
                "pm_name":      rows[0].pm_name,
            }

        latest_props = sorted({r.property_name for r in rtype_rows
                                if r.year == latest_yr and r.month == latest_mo})

        for prop_idx, prop_name in enumerate(latest_props):
            curr = _prop_agg(prop_name, latest_yr, latest_mo)
            if curr is None:
                continue
            # Alternating ice-blue fill
            row_fill = _ICE_BLUE_FILL if prop_idx % 2 == 0 else _WHITE_FILL
            for ci in range(1, len(prop_headers) + 1):
                ws.cell(row, ci).fill = row_fill
            ws.cell(row, 1, prop_name)
            ws.cell(row, 2, curr["pm_name"])
            _c(ws, row, 3, curr["current_owed"], CURRENCY_FMT)
            _c(ws, row, 4, curr["prepayments"],  CURRENCY_FMT)
            _c(ws, row, 5, curr["pct_overdue"],  PCT_FMT)

            if prior_period:
                prev = _prop_agg(prop_name, prior_period[0], prior_period[1])
                if prev:
                    co_delta = curr["current_owed"] - prev["current_owed"]
                    _c(ws, row, 6, co_delta, CURRENCY_FMT)
                    # No variance fill on $ amount YoY column — only % >60 Days gets shading
                    if curr["pct_overdue"] is not None and prev["pct_overdue"] is not None:
                        pct_delta = curr["pct_overdue"] - prev["pct_overdue"]
                        _c(ws, row, 7, pct_delta, PCT_FMT)
                        # >5pp threshold for % >60 Days YoY delta
                        apply_variance_fill(ws.cell(row, 7), pct_delta,
                                            favorable_is_positive=False, threshold=0.05)
            # Bad debt write-off from financial statements (col 8)
            bd_val = _bd.get((prop_name, latest_yr, latest_mo))
            if bd_val:
                _c(ws, row, 8, bd_val, CURRENCY_FMT)
            row += 1

        # Excel Table for property analysis block (no freeze panes on AR Aging tab)
        if row > data_start_row:
            last_prop_row = row - 1
            t_name = _next_table_name("PropertyAR_")
            tab = Table(
                displayName=t_name,
                ref=f"A{hdr_row}:{get_column_letter(len(prop_headers))}{last_prop_row}",
            )
            tab.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False, showLastColumn=False,
                showRowStripes=False, showColumnStripes=False,
            )
            ws.add_table(tab)

        row += 2  # blank rows between blocks

    # No freeze_panes on AR Aging tab — each block starts at a different row
    _autofit_columns(ws)


def _autofit_columns(ws, max_width=60):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                length = len(str(cell.value or ""))
                if length > max_len:
                    max_len = length
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, max_width)
