"""Builds the 8-tab backup/audit workbook (7 original + AR_Aging_Detail)."""

from collections import defaultdict
import openpyxl
from openpyxl.utils import get_column_letter
from app.models import (
    MappedRow, PropertyPeriodKPIs, SourceIndexEntry, MappingEntry, QualityCheck
)
from app.exporter.styles import (
    style_header_row, add_kpi_comment,
    CURRENCY_FMT, PCT_FMT, VAR_PCT_FMT, RED_FILL,
)
from config import ECO_OCC_TARGET


def build_backup_workbook(
    mapped_rows: list[MappedRow],
    kpis: list[PropertyPeriodKPIs],
    source_index: list[SourceIndexEntry],
    mapping_entries: list[MappingEntry],
    quality_checks: list[QualityCheck],
    output_path: str,
    eco_occ_target: float = ECO_OCC_TARGET,
    ar_rows: list | None = None,
) -> str:
    """Builds backup workbook at output_path. Returns path."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _build_raw_data(wb, mapped_rows)
    _build_source_index(wb, source_index)
    _build_assumptions_mapping(wb, mapping_entries)
    _build_budget_vs_actual(wb, kpis)
    _build_account_detail(wb, mapped_rows)
    _build_economic_occupancy(wb, kpis)
    _build_quality_checks(wb, quality_checks, kpis, eco_occ_target)
    _build_ar_aging_detail(wb, ar_rows or [])

    wb.save(output_path)
    return output_path


def _build_raw_data(wb, mapped_rows):
    ws = wb.create_sheet("Raw_Data")
    headers = [
        "Property", "Property Manager", "Source Workbook", "Source Sheet",
        "Source Type", "Source Row", "Account Code", "Account Name",
        "Account Category", "KPI Mapping",
        "Year", "Month", "Period", "Amount", "Original Amount", "Notes",
    ]
    _write_header(ws, headers, 1)
    for row_idx, r in enumerate(mapped_rows, 2):
        ws.cell(row_idx, 1,  r.property_name)
        ws.cell(row_idx, 2,  r.pm_name)
        ws.cell(row_idx, 3,  r.source_workbook)
        ws.cell(row_idx, 4,  r.source_sheet)
        ws.cell(row_idx, 5,  r.source_type)
        ws.cell(row_idx, 6,  r.source_row)
        ws.cell(row_idx, 7,  r.account_code)
        ws.cell(row_idx, 8,  r.account_name)
        ws.cell(row_idx, 9,  r.account_category)
        ws.cell(row_idx, 10, r.kpi_mapping)
        ws.cell(row_idx, 11, r.year)
        ws.cell(row_idx, 12, r.month)
        ws.cell(row_idx, 13, _month_name(r.month))
        ws.cell(row_idx, 14, r.amount);          ws.cell(row_idx, 14).number_format = CURRENCY_FMT
        ws.cell(row_idx, 15, r.original_amount); ws.cell(row_idx, 15).number_format = CURRENCY_FMT
        ws.cell(row_idx, 16, r.notes)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _build_source_index(wb, source_index):
    ws = wb.create_sheet("Source_Index")
    headers = [
        "Source Workbook", "Source Sheet", "Property", "Property Manager",
        "Year", "Source Type", "Processed?", "Rows Extracted",
        "Reason if Excluded", "Notes",
    ]
    _write_header(ws, headers, 1)
    for i, e in enumerate(source_index, 2):
        ws.cell(i, 1, e.source_workbook); ws.cell(i, 2, e.source_sheet)
        ws.cell(i, 3, e.property_name);   ws.cell(i, 4, e.pm_name)
        ws.cell(i, 5, e.year);            ws.cell(i, 6, e.source_type)
        ws.cell(i, 7, "Yes" if e.processed else "No")
        ws.cell(i, 8, e.rows_extracted)
        ws.cell(i, 9, e.reason_if_excluded)
        ws.cell(i, 10, e.notes)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _build_assumptions_mapping(wb, mapping_entries):
    ws = wb.create_sheet("Assumptions_Mapping")
    headers = [
        "Account Code", "Account Name", "Assigned Category", "KPI Mapping",
        "Treatment", "Include in NOI?", "Include in Eco Occ?", "Notes",
    ]
    _write_header(ws, headers, 1)
    for i, e in enumerate(mapping_entries, 2):
        ws.cell(i, 1, e.account_code); ws.cell(i, 2, e.account_name)
        ws.cell(i, 3, e.assigned_category); ws.cell(i, 4, e.kpi_mapping)
        ws.cell(i, 5, e.treatment)
        ws.cell(i, 6, "Yes" if e.include_in_noi else "No")
        ws.cell(i, 7, "Yes" if e.include_in_eco_occ else "No")
        ws.cell(i, 8, e.notes)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _build_budget_vs_actual(wb, kpis):
    ws = wb.create_sheet("Budget_vs_Actual")
    headers = [
        "Property", "Property Manager", "Total Units",
        "Year", "Month", "Period",
        "Actual Income", "Budget Income", "Income Variance", "Income Variance %",
        "Actual Expenses", "Budget Expenses", "Expense Variance", "Expense Variance %",
        "Actual NOI", "Budget NOI", "NOI Variance", "NOI Variance %",
        "Income/Unit", "Expense/Unit", "NOI/Unit",
    ]
    _write_header(ws, headers, 1)
    currency_cols = {7, 8, 9, 11, 12, 13, 15, 16, 17, 19, 20, 21}
    pct_cols      = {10, 14, 18}
    for i, k in enumerate(sorted(kpis, key=lambda x: (x.property_name, x.year, x.month)), 2):
        vals = [
            k.property_name, k.pm_name,
            k.total_units if k.total_units is not None else "N/A",
            k.year, k.month, k.period,
            k.actual_income, k.budget_income, k.income_variance, k.income_variance_pct,
            k.actual_expenses, k.budget_expenses, k.expense_variance, k.expense_variance_pct,
            k.actual_noi, k.budget_noi, k.noi_variance, k.noi_variance_pct,
            k.income_per_unit, k.expense_per_unit, k.noi_per_unit,
        ]
        for col, val in enumerate(vals, 1):
            ws.cell(i, col, val)
            if col in currency_cols:
                ws.cell(i, col).number_format = CURRENCY_FMT
            elif col in pct_cols:
                ws.cell(i, col).number_format = VAR_PCT_FMT
    ws.freeze_panes = "D2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _build_account_detail(wb, mapped_rows):
    ws = wb.create_sheet("Account_Detail")
    headers = [
        "Property", "Property Manager", "Year", "Month", "Period",
        "Account Code", "Account Name", "KPI Mapping",
        "Actual Amount", "Budget Amount", "Budget Variance",
        "Total Units", "Amount Per Unit", "Notes",
    ]
    _write_header(ws, headers, 1)

    act: dict = defaultdict(float)
    bud: dict = defaultdict(float)
    meta: dict = {}
    for r in mapped_rows:
        key = (r.property_name, r.pm_name, r.year, r.month, r.account_code, r.account_name, r.kpi_mapping)
        if r.source_type == "Budget":
            bud[key] += r.amount
        else:
            act[key] += r.amount
        meta[key] = r

    all_keys = set(act.keys()) | set(bud.keys())
    row = 2
    for key in sorted(all_keys):
        prop, pm, yr, mo, code, name, kpi = key
        actual_amt = act.get(key)
        budget_amt = bud.get(key)
        variance = (actual_amt - budget_amt) if (actual_amt is not None and budget_amt is not None) else None
        ws.cell(row, 1, prop); ws.cell(row, 2, pm)
        ws.cell(row, 3, yr);   ws.cell(row, 4, mo); ws.cell(row, 5, _month_name(mo))
        ws.cell(row, 6, code); ws.cell(row, 7, name); ws.cell(row, 8, kpi)
        ws.cell(row, 9, actual_amt);  ws.cell(row, 9).number_format = CURRENCY_FMT
        ws.cell(row, 10, budget_amt); ws.cell(row, 10).number_format = CURRENCY_FMT
        ws.cell(row, 11, variance);   ws.cell(row, 11).number_format = CURRENCY_FMT
        ws.cell(row, 12, "N/A")
        ws.cell(row, 13, "N/A")
        row += 1

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _build_economic_occupancy(wb, kpis):
    ws = wb.create_sheet("Economic_Occupancy")
    headers = [
        "Property", "Property Manager", "Year", "Month", "Period",
        "GPR / Rental Income", "Vacancy", "Concessions", "Bad Debt",
        "Net Collectible Rental Revenue", "Economic Occupancy %",
        "Budget Eco Occ %", "Eco Occ Variance",
        "Total Units",
    ]
    _write_header(ws, headers, 1)
    for i, k in enumerate(sorted(kpis, key=lambda x: (x.property_name, x.year, x.month)), 2):
        ws.cell(i, 1, k.property_name); ws.cell(i, 2, k.pm_name)
        ws.cell(i, 3, k.year);          ws.cell(i, 4, k.month); ws.cell(i, 5, k.period)
        ws.cell(i, 6, k.gpr);           ws.cell(i, 6).number_format = CURRENCY_FMT
        ws.cell(i, 7, k.vacancy);       ws.cell(i, 7).number_format = CURRENCY_FMT
        ws.cell(i, 8, k.concessions);   ws.cell(i, 8).number_format = CURRENCY_FMT
        ws.cell(i, 9, k.bad_debt);      ws.cell(i, 9).number_format = CURRENCY_FMT
        ws.cell(i, 10, k.net_collectible); ws.cell(i, 10).number_format = CURRENCY_FMT
        ws.cell(i, 11, k.eco_occ_pct);     ws.cell(i, 11).number_format = PCT_FMT
        ws.cell(i, 12, k.budget_eco_occ_pct); ws.cell(i, 12).number_format = PCT_FMT
        ws.cell(i, 13, k.eco_occ_variance);   ws.cell(i, 13).number_format = PCT_FMT
        ws.cell(i, 14, k.total_units if k.total_units is not None else "N/A")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _build_quality_checks(wb, quality_checks, kpis, eco_occ_target):
    ws = wb.create_sheet("Quality_Checks")
    _write_header(ws, ["Check", "Status", "Detail"], 1)
    row = 2
    for qc in quality_checks:
        ws.cell(row, 1, qc.check_name)
        ws.cell(row, 2, "PASS" if qc.passed else "FAIL")
        ws.cell(row, 3, qc.detail)
        if not qc.passed:
            ws.cell(row, 2).fill = RED_FILL
        row += 1
    for k in kpis:
        if k.eco_occ_pct is not None and (k.eco_occ_pct > 1.0 or k.eco_occ_pct < 0):
            ws.cell(row, 1, f"Eco Occ out of range: {k.property_name} {k.year}-{k.month:02d}")
            ws.cell(row, 2, "REVIEW")
            ws.cell(row, 3, f"Eco Occ = {k.eco_occ_pct:.1%}. Possible sign error in source data.")
            row += 1
    ws.freeze_panes = "A2"


def _write_header(ws, headers, row_num):
    for col, h in enumerate(headers, 1):
        ws.cell(row_num, col, h)
    style_header_row(ws, row_num, len(headers))
    ws.freeze_panes = f"A{row_num + 1}"


def _month_name(month: int) -> str:
    names = {1:"Jan",2:"Feb",3:"Mar",4:"Apr",5:"May",6:"Jun",
             7:"Jul",8:"Aug",9:"Sep",10:"Oct",11:"Nov",12:"Dec"}
    return names.get(month, str(month))


def _build_ar_aging_detail(wb, ar_rows: list) -> None:
    """AR_Aging_Detail tab: one row per ARAgingRow, all fields + computed columns."""
    ws = wb.create_sheet("AR_Aging_Detail")
    headers = [
        "Property", "PM", "Source File", "Receivable Type",
        "Year", "Month", "Period",
        "Charge Amount", "Current Owed",
        "0–30", "31–60", "61–90", "Over 90",
        "Pre-payments", "Total Over 60", "% >60 Days",
    ]
    _write_header(ws, headers, 1)

    if not ar_rows:
        ws.cell(2, 1, "No AR Aging data was uploaded for this analysis.")
        return

    # Sort: Receivable Type, Property Name, Year, Month
    sorted_rows = sorted(ar_rows, key=lambda r: (r.receivable_type, r.property_name, r.year, r.month))

    for i, r in enumerate(sorted_rows, 2):
        charge   = r.charge_amount
        over_60  = r.owed_61_90 + r.owed_over_90
        pct_ov   = (over_60 / charge) if charge and charge > 0 else None

        ws.cell(i, 1,  r.property_name)
        ws.cell(i, 2,  r.pm_name)
        ws.cell(i, 3,  r.source_file)
        ws.cell(i, 4,  r.receivable_type)
        ws.cell(i, 5,  r.year)
        ws.cell(i, 6,  r.month)
        ws.cell(i, 7,  _month_name(r.month))
        ws.cell(i, 8,  r.charge_amount);  ws.cell(i, 8).number_format  = CURRENCY_FMT
        ws.cell(i, 9,  r.current_owed);   ws.cell(i, 9).number_format  = CURRENCY_FMT
        ws.cell(i, 10, r.owed_0_30);      ws.cell(i, 10).number_format = CURRENCY_FMT
        ws.cell(i, 11, r.owed_31_60);     ws.cell(i, 11).number_format = CURRENCY_FMT
        ws.cell(i, 12, r.owed_61_90);     ws.cell(i, 12).number_format = CURRENCY_FMT
        ws.cell(i, 13, r.owed_over_90);   ws.cell(i, 13).number_format = CURRENCY_FMT
        ws.cell(i, 14, r.prepayments);    ws.cell(i, 14).number_format = CURRENCY_FMT
        ws.cell(i, 15, over_60);          ws.cell(i, 15).number_format = CURRENCY_FMT
        ws.cell(i, 16, pct_ov);           ws.cell(i, 16).number_format = PCT_FMT

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
