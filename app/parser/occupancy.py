"""Parses physical occupancy Excel reports into OccupancyRow records.

Supports three layouts:

  Narrow (original):  one row per property-month.
    Columns: Property | Year | Month | Occupied Units | Total Units

  Wide/Yardi:  one row per property, months as columns (values are percentages 0-100).
    Example header:  Name | Units | Sq Ft | Jan | Feb | ... | Dec
    Title rows contain the report year ("Month Year = 01/2025").

  ConAm Trend ("Occupancy Trend Report"):  one row per property, one column per
    month (values are percentages 0-100), with month headers as datetime objects.
    Row 1 contains "Occupancy Trend Report".
    Row 8 contains the date headers; col A = property name (with " -Yardi" suffix);
    col D = total units.  Year is encoded in each datetime header.
"""

import datetime as _dt
import re
from typing import Optional
import openpyxl
from app.models import OccupancyRow
from app.parser.utils import load_workbook_any_format

# ── Narrow-format column keywords ─────────────────────────────────────────────
_PROPERTY_KEYWORDS = ["property", "prop", "name", "asset"]
_YEAR_KEYWORDS     = ["year", "yr"]
_MONTH_KEYWORDS    = ["month", "mo", "period"]
_OCCUPIED_KEYWORDS = ["occupied", "occ unit", "occ. unit", "units occ"]
_TOTAL_KEYWORDS    = ["total unit", "tot unit", "total units", "# units", "unit count"]

# Canonical month abbreviations Jan=1 ... Dec=12
_MONTH_ABBRS = ["jan", "feb", "mar", "apr", "may", "jun",
                "jul", "aug", "sep", "oct", "nov", "dec"]

_YEAR_RE = re.compile(r'\b(20\d{2})\b')

# Header cells that identify a "property name" column in the wide Yardi format.
# Split into two tiers so that an explicit "Name" column is always preferred over a
# "Property" column (which in Yardi exports is typically the property *code*, not the
# display name — e.g. 'f66', 'monte01').
#
# Detection logic: scan all header columns, record the best tier-1 match; fall back to
# tier-2 only if no tier-1 match is found.
_WIDE_NAME_HEADERS_PRIORITY: frozenset[str] = frozenset({
    # These headers reliably contain the full display name
    "name",
    "property name", "project name", "asset name", "building name",
})
_WIDE_NAME_HEADERS_FALLBACK: frozenset[str] = frozenset({
    # These headers *may* contain the display name, but in Yardi 12-Month Occupancy
    # reports the "Property" column is a short code — only use as last resort
    "property", "project", "asset", "prop", "building",
})
# Union for any caller that still wants a single set (e.g. tests, narrow-format helpers)
_WIDE_NAME_HEADERS: frozenset[str] = _WIDE_NAME_HEADERS_PRIORITY | _WIDE_NAME_HEADERS_FALLBACK

# Header cells that identify a "total units" column in the wide Yardi format
_WIDE_UNITS_HEADERS: frozenset[str] = frozenset({
    "units", "total units", "total", "# units", "unit count",
    "num units", "no. of units", "no of units", "# of units",
})

# Values in a percentage cell that should be treated as "no data for this month"
_SKIP_PCT_VALUES: frozenset[str] = frozenset({
    "", "-", "--", "n/a", "na", "none", "null",
})

# Rows whose Name-column value matches these patterns are summary/filler rows to skip
_SKIP_NAME_PATTERNS: frozenset[str] = frozenset({
    "total", "portfolio total", "grand total", "portfolio", "subtotal",
})


def parse_occupancy_report(file_path: str) -> list[OccupancyRow]:
    wb = load_workbook_any_format(file_path)
    rows: list[OccupancyRow] = []

    for ws in wb.worksheets:
        all_cells = list(ws.iter_rows(values_only=True))

        # ConAm "Occupancy Trend Report" — wide with datetime column headers
        if _is_conam_trend_format(all_cells):
            trend_rows = _parse_conam_trend_format(all_cells)
            if trend_rows:
                rows.extend(trend_rows)
            continue

        # Yardi wide format (months as text abbreviation headers, values are percentages)
        wide_rows = _parse_wide_format(all_cells)
        if wide_rows:
            rows.extend(wide_rows)
            continue

        # Fall back to narrow format (one row per property-month)
        header_idx, col_map = _find_narrow_header(all_cells)
        if header_idx is None:
            continue

        for row in all_cells[header_idx + 1:]:
            try:
                prop  = _get(row, col_map, "property")
                year  = int(_get(row, col_map, "year") or 0)
                month = int(_get(row, col_map, "month") or 0)
                occ   = int(_get(row, col_map, "occupied") or 0)
                total = int(_get(row, col_map, "total") or 0)
            except (TypeError, ValueError):
                continue
            if not prop or not year or not month or not total:
                continue
            rows.append(OccupancyRow(
                property_name=str(prop).strip(),
                year=year, month=month,
                occupied_units=occ, total_units=total,
            ))

    wb.close()
    return rows


# ── ConAm Occupancy Trend Report parser ──────────────────────────────────────

def _is_conam_trend_format(all_cells: list) -> bool:
    """Return True if this sheet is a ConAm Occupancy Trend Report."""
    if not all_cells:
        return False
    return str(all_cells[0][0] or "").strip() == "Occupancy Trend Report"


def _parse_conam_trend_format(all_cells: list) -> list[OccupancyRow]:
    """Parse the ConAm Occupancy Trend Report (wide, percentage, datetime headers).

    Layout:
      Row 1:  "Occupancy Trend Report"
      Row 8:  datetime headers at varying column indices (end-of-month dates)
      Row 10: "Group <none>" section separator — skipped
      Row 11+: property data rows
        Col A (0): property name, e.g. "Auburn Park - Yardi"
        Col D (3): total units (integer)
        Month cols: percentage 0-100 (occupied unit percentage)
    """
    # Find the date-header row: first row within the first 15 that has >= 6 datetime values
    date_row_idx: Optional[int] = None
    date_row: Optional[tuple] = None
    for idx, row in enumerate(all_cells[:15]):
        dt_count = sum(1 for c in row if isinstance(c, (_dt.date, _dt.datetime)))
        if dt_count >= 6:
            date_row_idx = idx
            date_row = row
            break

    if date_row_idx is None or date_row is None:
        return []

    # Build {col_index: (year, month)} from the datetime header row
    col_to_ym: dict[int, tuple[int, int]] = {}
    for col_idx, cell in enumerate(date_row):
        if isinstance(cell, (_dt.date, _dt.datetime)):
            col_to_ym[col_idx] = (cell.year, cell.month)

    if not col_to_ym:
        return []

    # Total-units column — always col D (index 3) in this format
    total_units_col = 3

    rows: list[OccupancyRow] = []
    for row in all_cells[date_row_idx + 1:]:
        if not row or row[0] is None:
            continue

        name_raw = str(row[0]).strip()
        if not name_raw:
            continue

        # Skip section-header rows ("Group <none>", "Report Version:", etc.)
        if name_raw.lower().startswith(("group ", "report version")):
            continue

        # Strip " - Yardi", " -Yardi", " -yardi" suffixes (case-insensitive)
        name = re.sub(r'\s*-\s*Yardi\s*$', '', name_raw, flags=re.IGNORECASE).strip()
        if not name:
            continue

        # Total units
        total_cell = row[total_units_col] if total_units_col < len(row) else None
        try:
            total_units = int(round(float(str(total_cell).replace(',', '').rstrip('%'))))
        except (TypeError, ValueError):
            continue
        if total_units <= 0:
            continue

        # One OccupancyRow per month column
        for col_idx, (yr, mo) in col_to_ym.items():
            if col_idx >= len(row):
                continue
            pct_cell = row[col_idx]
            if pct_cell is None:
                continue
            pct_str = str(pct_cell).strip().rstrip('%').replace(',', '')
            if pct_str.lower() in _SKIP_PCT_VALUES:
                continue
            try:
                pct = float(pct_str)
            except (ValueError, AttributeError):
                continue
            if pct > 110:
                continue
            pct = max(0.0, pct)
            occupied_units = round(pct / 100 * total_units)
            rows.append(OccupancyRow(
                property_name=name,
                year=yr,
                month=mo,
                occupied_units=occupied_units,
                total_units=total_units,
            ))

    return rows


# ── Wide-format parser ─────────────────────────────────────────────────────────

def _parse_wide_format(all_cells: list) -> Optional[list[OccupancyRow]]:
    """Parse Yardi-style wide occupancy: months as columns, values are percentages 0-100.

    Returns a list of OccupancyRow records, or None if the sheet does not look
    like the wide format (allowing the caller to fall back to narrow format).

    Robustness notes:
    - Name column: accepts "Name", "Property", "Property Name", "Project", etc.
    - Units column: accepts "Units", "Total Units", "# Units", "Unit Count", etc.
    - Percentage values: accepts bare floats (95.0), strings ("95.0"), and
      percent-suffixed strings ("95.0%"). Skips cells that are blank, "--", "N/A".
    - Total units: accepts floats (rounded to int) and comma-formatted integers.
    - Summary rows: skips "Weighted Average", "Average", "Total", "Portfolio Total".
    """
    # Find the header row: must contain >= 6 month abbreviations as column headers.
    header_row_idx: Optional[int] = None
    header_row: Optional[tuple] = None
    for idx, row in enumerate(all_cells[:20]):
        row_lower = [str(c).lower().strip() if c else "" for c in row]
        month_hits = sum(1 for s in row_lower if s in _MONTH_ABBRS)
        if month_hits >= 6:
            header_row_idx = idx
            header_row = row
            break

    if header_row_idx is None or header_row is None:
        return None

    # Map month integer -> column index; also locate Name and Units columns.
    # Use two-tier name detection: prefer tier-1 ("Name", "Property Name", …) over
    # tier-2 ("Property", "Prop", …).  In Yardi 12-Month Occupancy reports the
    # "Property" column holds a short code (e.g. 'f66') while the "Name" column holds
    # the actual display name.  Taking the first match without priority would silently
    # return codes instead of names, breaking the downstream property-name join.
    month_cols: dict[int, int] = {}
    name_col_priority: Optional[int] = None   # tier-1: "name", "property name", …
    name_col_fallback: Optional[int] = None   # tier-2: "property", "prop", …
    units_col: Optional[int] = None

    for col_idx, cell in enumerate(header_row):
        if cell is None:
            continue
        s = str(cell).lower().strip()
        if s in _MONTH_ABBRS:
            month_cols[_MONTH_ABBRS.index(s) + 1] = col_idx
        elif name_col_priority is None and s in _WIDE_NAME_HEADERS_PRIORITY:
            name_col_priority = col_idx
        elif name_col_fallback is None and s in _WIDE_NAME_HEADERS_FALLBACK:
            name_col_fallback = col_idx
        elif units_col is None and s in _WIDE_UNITS_HEADERS:
            units_col = col_idx

    # Resolve name column: prefer tier-1 match, fall back to tier-2
    name_col: Optional[int] = name_col_priority if name_col_priority is not None else name_col_fallback

    # Need both a name column and a units column to proceed.
    if name_col is None or units_col is None:
        return None

    # Extract year from title rows above the header.
    year = 0
    for row in all_cells[:header_row_idx]:
        for cell in row:
            if cell is None:
                continue
            m = _YEAR_RE.search(str(cell))
            if m:
                year = int(m.group(1))
                break
        if year:
            break

    if year == 0:
        return None  # Cannot determine report year

    # Parse data rows.
    rows: list[OccupancyRow] = []
    for row in all_cells[header_row_idx + 1:]:
        if not row or all(c is None for c in row):
            continue

        # Property name from the "Name" column
        name_cell = row[name_col] if name_col < len(row) else None
        if not name_cell:
            continue
        name = str(name_cell).strip()
        if not name:
            continue

        # Skip summary / weighted-average / total rows
        name_lower = name.lower()
        if (name_lower in _SKIP_NAME_PATTERNS
                or "weighted average" in name_lower
                or name_lower.startswith("average")):
            continue

        # Total units: accept floats (e.g. "100.0") and comma-formatted strings
        units_cell = row[units_col] if units_col < len(row) else None
        try:
            units_str = str(units_cell).strip().replace(',', '').rstrip('%') if units_cell is not None else "0"
            total_units = int(round(float(units_str)))
        except (TypeError, ValueError):
            continue
        if total_units <= 0:
            continue

        # One OccupancyRow per month column
        for month, col_idx in month_cols.items():
            if col_idx >= len(row):
                continue
            pct_cell = row[col_idx]
            if pct_cell is None:
                continue

            # Parse percentage — handle bare float, "95.0", "95.0%", "N/A", "--"
            pct_str = str(pct_cell).strip().rstrip('%').replace(',', '')
            if pct_str.lower() in _SKIP_PCT_VALUES:
                continue
            try:
                pct = float(pct_str)
            except (ValueError, AttributeError):
                continue

            # Sanity-check: allow 0–110 to accommodate minor rounding artefacts.
            # Values slightly below 0 are treated as 0 (rounding); values > 110 are
            # likely data errors (e.g. a dollar amount was placed in a % column).
            if pct > 110:
                continue
            pct = max(0.0, pct)

            occupied_units = round(pct / 100 * total_units)
            rows.append(OccupancyRow(
                property_name=name,
                year=year,
                month=month,
                occupied_units=occupied_units,
                total_units=total_units,
            ))

    return rows if rows else None


# ── Narrow-format helpers ──────────────────────────────────────────────────────

def _find_narrow_header(all_cells: list) -> tuple:
    for idx, row in enumerate(all_cells[:20]):
        row_lower = [str(c).lower() if c else "" for c in row]
        col_map: dict[str, int] = {}
        for i, h in enumerate(row_lower):
            if any(k in h for k in _PROPERTY_KEYWORDS):
                col_map.setdefault("property", i)
            elif any(k in h for k in _YEAR_KEYWORDS):
                col_map.setdefault("year", i)
            elif any(k in h for k in _MONTH_KEYWORDS):
                col_map.setdefault("month", i)
            elif any(k in h for k in _OCCUPIED_KEYWORDS):
                col_map.setdefault("occupied", i)
            elif any(k in h for k in _TOTAL_KEYWORDS):
                col_map.setdefault("total", i)
        if all(k in col_map for k in ("property", "year", "month", "occupied", "total")):
            return idx, col_map
    return None, {}


def _get(row, col_map: dict, key: str):
    idx = col_map.get(key)
    if idx is None or idx >= len(row):
        return None
    return row[idx]
