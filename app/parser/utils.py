"""Shared parser utilities."""

import os
import tempfile
from typing import Optional
import openpyxl


def load_workbook_any_format(path: str):
    """Load an Excel file as an openpyxl workbook regardless of format.

    openpyxl only handles .xlsx/.xlsm.  For legacy .xls files (Excel 97-2003)
    we read with pandas/xlrd and write a temporary .xlsx so the rest of the
    parsing logic is completely unaffected.

    Raises ValueError with a clear message if the file is not a recognised
    Excel format, so the upload route can surface a friendly error to the user.
    """
    ext = os.path.splitext(path)[1].lower()

    if ext in (".xlsx", ".xlsm"):
        return openpyxl.load_workbook(path, data_only=True)

    if ext == ".xls":
        try:
            import pandas as pd
        except ImportError as exc:
            raise ImportError("pandas is required to read .xls files") from exc

        try:
            all_sheets = pd.read_excel(
                path, sheet_name=None, header=None, dtype=object, engine="xlrd"
            )
        except Exception as exc:
            raise ValueError(
                f"Could not read {os.path.basename(path)} as an .xls file: {exc}"
            ) from exc

        # Write to a temporary .xlsx so the existing parser logic runs unchanged.
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp_path = tmp.name
        tmp.close()
        try:
            with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
                for sheet_name, df in all_sheets.items():
                    # Truncate sheet names to 31 chars (Excel limit)
                    safe_name = str(sheet_name)[:31]
                    df.to_excel(writer, sheet_name=safe_name, header=False, index=False)
            return openpyxl.load_workbook(tmp_path, data_only=True)
        finally:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass

    raise ValueError(
        f"{os.path.basename(path)} is not a supported Excel format "
        f"(got '{ext}', expected .xlsx or .xls)."
    )
