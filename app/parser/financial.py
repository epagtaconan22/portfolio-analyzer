"""Parses financial statement Excel workbooks into RawRow records."""

import datetime as _dt
import itertools
import os
import re
from collections import Counter
from typing import Optional
import openpyxl
from app.models import RawRow, SourceIndexEntry
from app.parser.sheet_inferrer import (
    infer_sheet_type,
    _ACTUAL_BUDGET_PATTERNS,
    _ACTUAL_PATTERNS,
    _BUDGET_PATTERNS,
)

# Short 3-letter abbreviations are sufficient — substring matching means "jan" already
# matches "january", "jun" matches "june", etc.  Long-form duplicates are not needed.
_MONTH_PATTERNS = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}

# Compiled once for use in _find_header_row — avoids per-cell linear key scan.
_MONTH_RE = re.compile(r'\b(?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)\b')

# Matches MM/YYYY, MM-YYYY, YYYY/MM, YYYY-MM (numeric date column headers).
_NUMERIC_MONTH_RE = re.compile(
    r'\b(?:0?[1-9]|1[0-2])[/\-](?:20\d{2})\b'   # MM/YYYY or MM-YYYY
    r'|'
    r'\b(?:20\d{2})[/\-](?:0?[1-9]|1[0-2])\b'   # YYYY/MM or YYYY-MM
)

# Matches MM/YY or MM-YY (2-digit year — JSCO format: "01/25", "12/26", etc.)
# The negative lookahead (?!\d) prevents matching MM/YYYY already caught above.
_NUMERIC_MONTH_YY_RE = re.compile(r'\b(0?[1-9]|1[0-2])[/\-](\d{2})(?!\d)\b')

# Keywords to strip from sheet names when inferring a property name, built from the
# same source as sheet_inferrer so the two modules stay in sync.
_SHEET_NAME_STRIP_PATTERNS = _ACTUAL_BUDGET_PATTERNS + _ACTUAL_PATTERNS + _BUDGET_PATTERNS
_SHEET_NAME_STRIP_RE = re.compile(
    "|".join(re.escape(p) for p in sorted(_SHEET_NAME_STRIP_PATTERNS, key=len, reverse=True)),
    flags=re.IGNORECASE,
)

# Account code: starts with a digit, followed by 2+ digits or common separators.
# Also matches letter-prefix codes like MR5120000 (JSCO: 1-3 uppercase letters + 4+ digits).
_ACCOUNT_CODE_RE = re.compile(r'^\d[\d\-\.:/]{2,}$|^[A-Z]{1,3}\d{4,}$')

_SKIP_ACCOUNT_PATTERNS = {
    "total", "subtotal", "net income", "net loss", "total income",
    "total revenue", "total expenses", "total cost", "total operating",
    "grand total", "net operating",
}

_YEAR_RE = re.compile(r'\b(20\d{2})\b')

# JSCO sheet names — these sheets use a different layout (header at row 8-9,
# MR-prefix account codes, MM/YY date format, property name in col D).
_JSCO_SHEET_NAMES = {"jsc_actual12", "leo_forecast"}

# Fragments that identify metadata/boilerplate rows in JSCO sheets — used to
# filter out non-property-name content when searching col D for a property name.
_JSCO_METADATA_FRAGMENTS = (
    "database", "entity", "page:", "date:", "time:", "accrual",
    "john stewart", "johnstewart", "12 month", "company",
)


def parse_financial_workbooks(
    file_paths: list[str],
    pm_name_map: dict[str, str],   # filename (basename) → PM company name
) -> tuple[list[RawRow], list[SourceIndexEntry]]:
    all_rows: list[RawRow] = []
    source_index: list[SourceIndexEntry] = []

    for path in file_paths:
        fname = os.path.basename(path)
        pm_name = pm_name_map.get(fname, _infer_pm_from_filename(fname))
        wb = openpyxl.load_workbook(path, data_only=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows, entry = _parse_sheet(ws, sheet_name, fname, pm_name)
            all_rows.extend(rows)
            source_index.append(entry)

    return all_rows, source_index


def _make_index_entry(
    source_workbook: str,
    sheet_name: str,
    pm_name: str,
    *,
    property_name: str = "",
    year: int = 0,
    source_type: str = "Unknown",
    processed: bool = False,
    rows_extracted: int = 0,
    reason_if_excluded: str = "",
) -> SourceIndexEntry:
    """Single construction point for SourceIndexEntry to avoid copy-paste."""
    return SourceIndexEntry(
        source_workbook=source_workbook,
        source_sheet=sheet_name,
        property_name=property_name,
        pm_name=pm_name,
        year=year,
        source_type=source_type,
        processed=processed,
        rows_extracted=rows_extracted,
        reason_if_excluded=reason_if_excluded,
    )


def _parse_sheet(
    ws, sheet_name: str, source_workbook: str, pm_name: str
) -> tuple[list[RawRow], SourceIndexEntry]:
    all_cells = list(ws.iter_rows(values_only=True))
    if not all_cells:
        return [], _make_index_entry(
            source_workbook, sheet_name, pm_name,
            reason_if_excluded="Empty sheet",
        )

    is_jsco = sheet_name.lower() in _JSCO_SHEET_NAMES

    title_rows = [list(r) for r in all_cells[:5]]
    header_row_idx, header_row = _find_header_row(all_cells)
    if header_row_idx is None:
        return [], _make_index_entry(
            source_workbook, sheet_name, pm_name,
            reason_if_excluded="No month header row found",
        )

    header_strs = [_cell_to_str(c) for c in header_row]
    source_type = infer_sheet_type(sheet_name, header_strs, title_rows)
    property_name = _infer_property_name(sheet_name, title_rows, is_jsco=is_jsco)

    col_map = _map_month_columns(header_strs, source_type)
    if not col_map:
        return [], _make_index_entry(
            source_workbook, sheet_name, pm_name,
            property_name=property_name, year=0, source_type=source_type,
            reason_if_excluded="No month columns identified",
        )

    # Derive sheet-level year from the most common per-column year.
    col_years = [yr for (_, _, yr) in col_map.values() if yr > 0]
    if col_years:
        year = Counter(col_years).most_common(1)[0][0]
    else:
        year = _infer_year(sheet_name, title_rows, header_strs)

    # JSCO trailing-12M: only keep the highest year's columns (Option A).
    # A single JSCO file may span two calendar years (e.g. Apr-2025 → Mar-2026).
    # To prevent double-counting when both the 2025 and 2026 files are uploaded,
    # retain only the most-recent year's columns from any cross-year JSCO file.
    if is_jsco and col_years:
        unique_years = set(col_years)
        if len(unique_years) > 1:
            max_year = max(unique_years)
            col_map = {ci: (mo, st, yr) for ci, (mo, st, yr) in col_map.items()
                       if yr == max_year}
            year = max_year

    rows: list[RawRow] = []
    for row_idx, row in enumerate(all_cells[header_row_idx + 1:], start=header_row_idx + 2):
        account_code, account_name = _extract_account(row)
        if not account_name or _is_skip_row(account_name, account_code):
            continue
        # JSCO sheets: every valid data row has an MR-prefix account code.
        # Rows without a code are metadata, subtotals, or blank separators — skip them.
        if is_jsco and not account_code:
            continue

        for col_idx, (month, stream, col_year) in col_map.items():
            if col_idx >= len(row):
                continue
            raw_val = row[col_idx]
            if raw_val is None:
                continue
            try:
                amount = float(raw_val)
            except (ValueError, TypeError):
                continue

            row_year = col_year if col_year else year
            rows.append(RawRow(
                property_name=property_name,
                pm_name=pm_name,
                source_workbook=source_workbook,
                source_sheet=sheet_name,
                source_type=stream,
                source_row=row_idx,
                account_code=account_code,
                account_name=account_name,
                year=row_year,
                month=month,
                amount=amount,
                original_amount=amount,
            ))

    return rows, _make_index_entry(
        source_workbook, sheet_name, pm_name,
        property_name=property_name, year=year, source_type=source_type,
        processed=True, rows_extracted=len(rows),
    )


def _find_header_row(all_cells) -> tuple[Optional[int], Optional[tuple]]:
    """Return the first row (within the first 30) that contains 3+ month indicators.

    Handles text abbreviations ("Jan", "February"), numeric date headers ("1/2024",
    "2024-01", "01/25"), and Excel date values returned by openpyxl as datetime objects.
    """
    for idx, row in enumerate(all_cells[:30]):
        hits = 0
        for c in row:
            if c is None:
                continue
            if isinstance(c, (_dt.date, _dt.datetime)):
                hits += 1
                continue
            s = str(c).lower()
            if (_MONTH_RE.search(s) or _NUMERIC_MONTH_RE.search(s)
                    or _NUMERIC_MONTH_YY_RE.search(s)):
                hits += 1
        if hits >= 3:
            return idx, row
    return None, None


def _extract_month_and_year(s: str) -> Optional[tuple[int, int]]:
    """Return (month, year) from a column header string, or None if not a date column.

    Handles:
      - Text month names: "Jan", "February", "Jan 2026"
      - MM/YYYY or MM-YYYY: "01/2026", "1-2026"
      - YYYY/MM or YYYY-MM: "2026/01", "2026-01"
      - MM/YY  or MM-YY  : "01/26", "01-26"  (JSCO format — 2-digit year → 20xx)
    """
    s_lower = s.lower().strip()
    # MM/YYYY or MM-YYYY (4-digit year — highest priority to avoid ambiguity)
    m = re.search(r'\b(0?[1-9]|1[0-2])[/\-](20\d{2})\b', s_lower)
    if m:
        return int(m.group(1)), int(m.group(2))
    # YYYY/MM or YYYY-MM
    m = re.search(r'\b(20\d{2})[/\-](0?[1-9]|1[0-2])\b', s_lower)
    if m:
        return int(m.group(2)), int(m.group(1))
    # MM/YY (JSCO format — 2-digit year, assumes 21st century)
    m = _NUMERIC_MONTH_YY_RE.search(s_lower)
    if m:
        return int(m.group(1)), 2000 + int(m.group(2))
    # Text month name — extract year from same string if present ("Jan 2026")
    m_name = _MONTH_RE.search(s_lower)
    if m_name:
        month = _MONTH_PATTERNS[m_name.group()]
        m_year = _YEAR_RE.search(s_lower)
        year = int(m_year.group(1)) if m_year else 0
        return month, year
    return None


def _extract_month(s: str) -> Optional[int]:
    """Thin wrapper — returns only the month integer."""
    result = _extract_month_and_year(s)
    return result[0] if result else None


def _map_month_columns(
    header_strs: list[str], source_type: str
) -> dict[int, tuple[int, str, int]]:
    """Return {col_index: (month_int, stream, year)} where year=0 means unknown.

    For Actual+Budget sheets the stream is determined per-column by whether the header
    contains a budget keyword; for pure-Budget sheets every month column is 'Budget';
    for all other types every month column is 'Actual'.
    """
    # Resolve the stream-assignment strategy once, outside the column loop.
    if source_type == "Actual+Budget":
        def _stream(h_lower: str) -> str:
            return "Budget" if ("bud" in h_lower) else "Actual"
    elif source_type == "Budget":
        def _stream(h_lower: str) -> str:  # noqa: F811
            return "Budget"
    else:
        def _stream(h_lower: str) -> str:  # noqa: F811
            return "Actual"

    col_map: dict[int, tuple[int, str, int]] = {}
    for idx, h in enumerate(header_strs):
        h_lower = h.lower().strip()
        result = _extract_month_and_year(h_lower)
        if result is None:
            continue
        month, col_year = result
        col_map[idx] = (month, _stream(h_lower), col_year)
    return col_map


def _cell_to_str(c) -> str:
    """Convert a cell value to a string suitable for month/year detection.

    datetime objects are formatted as "Jan 2024" so _MONTH_RE and _YEAR_RE
    can match them without special-casing every downstream function.
    """
    if c is None:
        return ""
    if isinstance(c, (_dt.date, _dt.datetime)):
        return c.strftime("%b %Y")   # e.g. "Jan 2024"
    return str(c)


def _extract_account(row) -> tuple[str, str]:
    """Return (account_code, account_name) from the first 1-4 cells of a data row."""
    code = ""
    name = ""
    for cell in row[:4]:
        if cell is None:
            continue
        val = str(cell).strip()
        if not val:
            continue
        if not name and _ACCOUNT_CODE_RE.match(val):
            code = val
        elif not name:
            name = val
        elif not code and _ACCOUNT_CODE_RE.match(val):
            code = val
        else:
            break
    return code, name


def _is_skip_row(account_name: str, account_code: str = "") -> bool:
    lower = account_name.lower().strip()
    if lower in _SKIP_ACCOUNT_PATTERNS:
        return True
    # Skip Yardi-style subtotal rows whose names begin with "Total " or "Net "
    # (e.g. "Total Rent Revenue", "Net Rental Revenue", "NET OPERATING INCOME (Loss)").
    # Individual line-item accounts almost never start with these words.
    if lower.startswith(("total ", "net ")):
        return True
    # Yardi subtotal accounts: codes ending in -1798 are computed range subtotals
    # (e.g. ConAm's "BASE SCHEDULED RENT" at 41000-1798 = sum of 41000-1000 + 41000-1100).
    if account_code and account_code.endswith("-1798"):
        return True
    return False


# Matches Yardi title-row pattern: "Property Name (short_code)" where the
# parenthetical is 2–15 alphanumeric/hyphen/underscore characters.
_YARDI_TITLE_RE = re.compile(r'^(.+?)\s*\([A-Za-z0-9_\-]{2,15}\)\s*$')


def _infer_property_name(
    sheet_name: str, title_rows: list, *, is_jsco: bool = False
) -> str:
    """Extract property name, preferring Yardi-style title row or JSCO col D.

    For standard Yardi sheets: looks for "Property Name (code)" in col A of the
    first two title rows.
    For JSCO sheets: looks for the property name in col D (index 3) of the first
    five title rows — JSCO places metadata in col A and the property name further right.
    Falls back to sheet-name stripping for both formats.
    """
    # 1. Yardi "Name (code)" pattern in col A (non-JSCO only).
    if not is_jsco:
        for row in title_rows[:2]:
            if not row:
                continue
            cell = row[0]
            if not cell:
                continue
            val = str(cell).strip()
            m = _YARDI_TITLE_RE.match(val)
            if m:
                return m.group(1).strip()

    # 2. JSCO: property name in col D (index 3), cols D–G (indices 3-6).
    if is_jsco:
        for row in title_rows[:5]:
            if not row or len(row) <= 3:
                continue
            for cell in row[3:7]:
                if not cell:
                    continue
                val = str(cell).strip()
                val_lower = val.lower()
                if (len(val) >= 6
                        and " " in val
                        and not any(frag in val_lower for frag in _JSCO_METADATA_FRAGMENTS)
                        and not re.search(r'\d', val)
                        and ":" not in val
                        and re.match(r"^[A-Za-z\s\-\.\&\']+$", val)):
                    return val

    # 3. Sheet-name stripping (covers "Actual - Sunrise Apts" etc.).
    name = _SHEET_NAME_STRIP_RE.sub(" ", sheet_name).strip(" -–|").strip()
    # Reject short or single-word remnants (e.g. "JSC_12" after stripping).
    if name and len(name) >= 5 and " " in name:
        return name

    # 4. Last resort: first non-empty cell in the title area.
    for row in title_rows[:3]:
        for cell in row[:3]:
            if cell:
                return str(cell).strip()
    return "Unknown Property"


def _infer_year(sheet_name: str, title_rows: list, header_strs: list[str]) -> int:
    """Extract the first four-digit year (20xx) from sheet name, title rows, or headers."""
    title_cells = (str(c) for row in title_rows[:3] for c in row if c)
    for src in itertools.chain([sheet_name], title_cells, header_strs):
        m = _YEAR_RE.search(src)
        if m:
            return int(m.group(1))
    return 0


# Known PM company names — searched in filename (case-insensitive) before falling back
# to the raw stem. Add new names here as new PM companies are onboarded.
_KNOWN_PM_NAMES = ["Solari", "ConAm", "JSCO"]


def _infer_pm_from_filename(filename: str) -> str:
    fname_lower = filename.lower()
    for pm in _KNOWN_PM_NAMES:
        if pm.lower() in fname_lower:
            return pm
    return os.path.splitext(filename)[0].replace("_", " ").replace("-", " ").strip()
