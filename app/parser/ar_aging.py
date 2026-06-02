"""Parses AR Aging export files into ARAgingRow records.

Three formats are supported, detected automatically from workbook structure:

ConAm 2026+ (conam_2026)
  Single sheet "Report1", all properties on one tab.
  Row 3:  "Trans through: MM/YYYY"
  Row 4:  (None, None, "Total", ...)
  Rows 5-6: column headers (Charge Code / Description / Total Unpaid Charges /
            0-30 / 31-60 / 61-90 / Over 90 / Prepays / Suspense / Balance)
  Data rows: property name rows have no numeric data in cols C-J.
  Tenant Rent  = charge codes that start with "RENT"  (captures RENT + RENTPM).
  Subsidy      = charge codes that start with "SUBRNT" (captures SUBRNT + SUBRNTPM).
  Amounts summed per (property, receivable type) within the file.

ConAm 2025 and prior (conam_2025)
  Multiple sheets, one per property (sheet name = short property name).
  Row 4 of each sheet contains a multi-line filter-criteria string with
    "As of MM/DD/YY".
  Section "Community Totals By Balance Type:" is found by scanning col B.
  Aging buckets at fixed Excel column indices (0-based):
    22=Current (0-30), 25=Over 30 (31-60), 28=Over 60 (61-90),
    31=Over 90, 34=Over 120  (Over 90 + Over 120 → owed_over_90),
    36=Balance (current_owed and charge_amount denominator).
  Balance type column: index 17 (Excel col R).
  Tenant Rent = balance type exactly "Rent -" (case-insensitive).
  Subsidy     = balance type starting with "Subsidy -" (case-insensitive).

Legacy / Solari (legacy)
  Single sheet, filename encodes PM + receivable type + period.
  E.g. "Solari_AR Aging_Tenant Rent_03_2024.xlsx".
  Row-per-property data starting after 5 header rows.
"""

import os
import re
from typing import Optional
import openpyxl

from app.models import ARAgingRow

_TYPE_NORMALIZE: dict[str, str] = {
    "tenant rent":        "Tenant Rent",
    "tenant receivable":  "Tenant Rent",
    "subsidy":            "Subsidy",
    "subsidy receivable": "Subsidy",
}


def parse_ar_aging_reports(file_paths: list[str]) -> list[ARAgingRow]:
    """Parse one or more AR Aging files. Returns combined list of ARAgingRow."""
    results: list[ARAgingRow] = []
    for path in file_paths:
        results.extend(_parse_one(path))
    return results


def _parse_one(path: str) -> list[ARAgingRow]:
    fname = os.path.basename(path)
    stem  = os.path.splitext(fname)[0]
    pm_name, receivable_type, year, month = _parse_filename(stem)

    wb  = openpyxl.load_workbook(path, data_only=True)
    fmt = _detect_format(wb)

    if fmt == "conam_2026":
        return _parse_conam_2026(wb, fname, pm_name or "ConAm", year, month)
    if fmt == "conam_2025":
        return _parse_conam_2025(wb, fname, pm_name or "ConAm", year, month)
    return _parse_legacy(wb, fname, pm_name, receivable_type, year, month)


# ─── Format detection ─────────────────────────────────────────────────────────

def _detect_format(wb) -> str:
    """Return 'conam_2026', 'conam_2025', or 'legacy'."""
    # ConAm 2026+: single sheet, row 3 starts with "Trans through"
    if len(wb.sheetnames) == 1:
        ws   = wb.worksheets[0]
        rows = list(ws.iter_rows(values_only=True, max_row=4))
        if len(rows) >= 3:
            cell = str(rows[2][0] or "").strip()
            if cell.lower().startswith("trans through"):
                return "conam_2026"
    # ConAm 2025: multiple sheets, first sheet row 1 = "A/R Aging Report"
    if len(wb.sheetnames) > 1:
        ws   = wb.worksheets[0]
        rows = list(ws.iter_rows(values_only=True, max_row=2))
        if rows and str(rows[0][0] or "").strip() == "A/R Aging Report":
            return "conam_2025"
    return "legacy"


# ─── ConAm 2026+ parser ───────────────────────────────────────────────────────

def _parse_conam_2026(wb, fname, pm_name, year, month) -> list[ARAgingRow]:
    """All-properties-on-one-tab format used from 2026 onward."""
    ws       = wb.worksheets[0]
    all_rows = list(ws.iter_rows(values_only=True))

    if year is None or month is None:
        year, month = _parse_period_trans_through(all_rows)

    # Accumulate per (property_name, receivable_type) so RENT + RENTPM sum together.
    acc: dict[tuple[str, str], dict] = {}
    current_prop: Optional[str] = None

    for row in all_rows[6:]:          # rows 1-6 are title / column headers
        col_a = row[0]
        if col_a is None:
            continue
        cell = str(col_a).strip()
        if not cell:
            continue
        if cell.lower().startswith("grand total"):
            break

        # Property header row: no numeric data in cols C–J (indices 2–8)
        is_property_header = (row[2] is None and row[3] is None and row[4] is None)
        if is_property_header:
            name = re.sub(r'\s*\([^)]+\)\s*$', '', cell).strip()
            current_prop = _fix_inverted_name(name)
            continue

        if current_prop is None:
            continue
        code = cell.upper()
        if code == "TOTAL":
            continue

        if code.startswith("RENT"):
            rtype = "Tenant Rent"
        elif code.startswith("SUBRNT"):
            rtype = "Subsidy"
        else:
            continue        # skip non-rent charge codes

        key = (current_prop, rtype)
        if key not in acc:
            acc[key] = dict(charge_amount=0.0, owed_0_30=0.0, owed_31_60=0.0,
                            owed_61_90=0.0, owed_over_90=0.0, prepayments=0.0,
                            current_owed=0.0)
        acc[key]["charge_amount"] += _f(row[2])   # Total Unpaid Charges
        acc[key]["owed_0_30"]     += _f(row[3])
        acc[key]["owed_31_60"]    += _f(row[4])
        acc[key]["owed_61_90"]    += _f(row[5])
        acc[key]["owed_over_90"]  += _f(row[6])
        acc[key]["prepayments"]   += _f(row[7])
        acc[key]["current_owed"]  += _f(row[9])   # Balance (net after prepays/suspense)

    return [
        ARAgingRow(
            property_name=prop, pm_name=pm_name, source_file=fname,
            receivable_type=rtype,
            year=year or 0, month=month or 0,
            **totals,
        )
        for (prop, rtype), totals in acc.items()
    ]


# ─── ConAm 2025 and prior parser ─────────────────────────────────────────────

def _parse_conam_2025(wb, fname, pm_name, year, month) -> list[ARAgingRow]:
    """One-sheet-per-property format used through 2025."""
    results: list[ARAgingRow] = []

    for sheet_name in wb.sheetnames:
        ws       = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))

        yr, mo = year, month
        if yr is None or mo is None:
            yr, mo = _parse_period_as_of(all_rows)

        prop_name = sheet_name.strip()   # PROPERTY_NAME_MAP normalises later

        # Locate "Community Totals By Balance Type:" section header
        bt_start = None
        for i, row in enumerate(all_rows):
            if row[1] is not None and "Community Totals By Balance Type" in str(row[1]):
                bt_start = i
                break
        if bt_start is None:
            continue

        # Accumulate rent and subsidy balance-type rows
        acc: dict[str, dict] = {}

        for row in all_rows[bt_start + 1:]:
            # Stop at the next major section header
            if row[1] is not None and "Community Totals By" in str(row[1]):
                break

            bt_cell = row[17]   # Excel column R (0-based index 17)
            if bt_cell is None:
                continue
            bt = str(bt_cell).strip()

            # Match only the plain rent and subsidy balance types.
            # "Rent -"         → Tenant Rent   (the form in the file is "Rent -")
            # "Subsidy - ..."  → Subsidy        (authority name follows the dash)
            # Exclude: "Rent Concession -", "Unapplied …", "Late Charge -", etc.
            bt_lower = bt.lower()
            if re.match(r'^rent\s*-\s*$', bt_lower):
                rtype = "Tenant Rent"
            elif bt_lower.startswith("subsidy -"):
                rtype = "Subsidy"
            else:
                continue

            if rtype not in acc:
                acc[rtype] = dict(charge_amount=0.0, owed_0_30=0.0, owed_31_60=0.0,
                                  owed_61_90=0.0, owed_over_90=0.0, prepayments=0.0,
                                  current_owed=0.0)

            current  = _f(row[22])   # Col W  — Current  (0–30 days)
            over_30  = _f(row[25])   # Col Z  — Over 30  (31–60 days)
            over_60  = _f(row[28])   # Col AC — Over 60  (61–90 days)
            over_90  = _f(row[31])   # Col AF — Over 90
            over_120 = _f(row[34])   # Col AI — Over 120 (merged into owed_over_90)
            balance  = _f(row[36])   # Col AK — Balance

            acc[rtype]["owed_0_30"]    += current
            acc[rtype]["owed_31_60"]   += over_30
            acc[rtype]["owed_61_90"]   += over_60
            acc[rtype]["owed_over_90"] += over_90 + over_120
            acc[rtype]["current_owed"] += balance
            # Use balance as the charge_amount denominator (no gross-charge column in this format)
            acc[rtype]["charge_amount"] += balance

        for rtype, totals in acc.items():
            results.append(ARAgingRow(
                property_name=prop_name, pm_name=pm_name, source_file=fname,
                receivable_type=rtype,
                year=yr or 0, month=mo or 0,
                **totals,
            ))

    return results


# ─── Legacy / Solari parser ───────────────────────────────────────────────────

def _parse_legacy(wb, fname, pm_name, receivable_type, year, month) -> list[ARAgingRow]:
    """Original row-per-property format (Solari and other PM companies)."""
    ws       = wb["Report1"] if "Report1" in wb.sheetnames else wb.worksheets[0]
    all_rows = list(ws.iter_rows(values_only=True))

    if year is None or month is None:
        year, month = _parse_period_from_sheet(all_rows)
    if receivable_type is None:
        receivable_type = _infer_type_from_stem(os.path.splitext(fname)[0])

    rows: list[ARAgingRow] = []
    for raw_row in all_rows[5:]:
        col_a = raw_row[0]
        if col_a is None:
            break
        raw_name = str(col_a).strip()
        if not raw_name or raw_name.startswith("Grand Total"):
            break
        property_name = re.sub(r'\s*\([^)]+\)\s*$', '', raw_name).strip()
        if not property_name:
            continue
        rows.append(ARAgingRow(
            property_name=property_name,
            pm_name=pm_name or "",
            source_file=fname,
            receivable_type=receivable_type or "Unknown",
            year=year or 0,
            month=month or 0,
            charge_amount=_f(raw_row[1]),
            current_owed=_f(raw_row[2]),
            owed_0_30=_f(raw_row[3]),
            owed_31_60=_f(raw_row[4]),
            owed_61_90=_f(raw_row[5]),
            owed_over_90=_f(raw_row[6]),
            prepayments=_f(raw_row[7]),
        ))
    return rows


# ─── Period helpers ───────────────────────────────────────────────────────────

def _parse_period_trans_through(all_rows) -> tuple[Optional[int], Optional[int]]:
    """ConAm 2026+: row 3 'Trans through: 03/2026'."""
    if len(all_rows) < 3 or all_rows[2][0] is None:
        return None, None
    m = re.search(r'(\d{1,2})[/\-](\d{4})', str(all_rows[2][0]))
    if m:
        return int(m.group(2)), int(m.group(1))
    return None, None


def _parse_period_as_of(all_rows) -> tuple[Optional[int], Optional[int]]:
    """ConAm 2025: row 4 multi-line cell 'As of 12/31/25'."""
    if len(all_rows) < 4 or all_rows[3][0] is None:
        return None, None
    m = re.search(r'As of (\d{1,2})/\d{1,2}/(\d{2,4})', str(all_rows[3][0]))
    if m:
        yr = int(m.group(2))
        if yr < 100:
            yr += 2000
        return yr, int(m.group(1))
    return None, None


def _parse_period_from_sheet(all_rows) -> tuple[Optional[int], Optional[int]]:
    """Legacy fallback: row 3 'Post To(MM/YY): 03/2024'."""
    if len(all_rows) < 3 or all_rows[2][0] is None:
        return None, None
    m = re.search(r'(\d{1,2})/(\d{4})', str(all_rows[2][0]))
    if m:
        return int(m.group(2)), int(m.group(1))
    return None, None


# ─── Filename parsing ─────────────────────────────────────────────────────────

def _parse_filename(stem: str) -> tuple[Optional[str], Optional[str], Optional[int], Optional[int]]:
    """
    Parse stem like "Solari_AR Aging_Tenant Rent_03_2024".
    Returns (pm_name, receivable_type, year, month).
    """
    parts = stem.split("_")
    if len(parts) < 4:
        return None, None, None, None
    try:
        month = int(parts[-2])
        year  = int(parts[-1])
    except ValueError:
        return None, None, None, None
    if not (1 <= month <= 12 and 2000 <= year <= 2100):
        return None, None, None, None

    pm_name  = parts[0]
    type_raw = " ".join(parts[2:-2]).lower().strip()
    receivable_type = _TYPE_NORMALIZE.get(type_raw)
    return pm_name, receivable_type, year, month


def _infer_type_from_stem(stem: str) -> str:
    stem_lower = stem.lower()
    if "subsidy" in stem_lower:
        return "Subsidy"
    return "Tenant Rent"


# ─── Utilities ────────────────────────────────────────────────────────────────

def _fix_inverted_name(name: str) -> str:
    """Convert 'Remi Apartments, The' → 'The Remi Apartments'."""
    m = re.match(r'^(.+),\s+The$', name, re.IGNORECASE)
    if m:
        return f"The {m.group(1).strip()}"
    return name


def _f(val) -> float:
    """Convert a cell value to float; treat None / non-numeric as 0.0."""
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


# Keep the old alias so existing call-sites that import _to_float continue to work.
_to_float = _f
